"""UNKNOWN 재정수단 상위 사업의 사람 검수 워크북 생성·검증."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_GUIDE = "작성안내"
SHEET_PROJECTS = "사업검수"
SHEET_YEARS = "연도별확인"
SHEET_CODES = "코드값"
SHEET_META = "원본메타"
HEADER_ROW = 4
DATA_START_ROW = 5

SCOPE_STATUSES = ("IN_SCOPE", "OUT_OF_SCOPE", "REVIEW_REQUIRED")
SCOPE_EXCLUSION_REASONS = (
    "INTERNAL_TRANSFER",
    "FINANCIAL_ASSET_OPERATION",
    "DEBT_REPAYMENT",
    "RESERVE_OR_SURPLUS_MANAGEMENT",
    "NON_POLICY_ADMINISTRATION",
    "OUTSIDE_TARGET_POLICY_SCOPE",
    "OTHER",
    "NOT_APPLICABLE",
    "REVIEW_REQUIRED",
)
APPLICABILITY_VALUES = ("APPLICABLE", "NOT_APPLICABLE", "REVIEW_REQUIRED")
FISCAL_INSTRUMENTS = (
    "DIRECT",
    "SUBSIDY",
    "CONTRIBUTION",
    "LOAN",
    "GUARANTEE",
    "EQUITY",
    "INTEREST_SUBSIDY",
    "RND",
    "FACILITY",
    "INFORMATIZATION",
    "OPERATION",
    "OTHER",
    "UNKNOWN",
)
YES_NO_REVIEW = ("YES", "NO", "REVIEW_REQUIRED")
CONFIDENCE_VALUES = ("HIGH", "MEDIUM", "LOW")
REVIEW_STATUSES = ("UNREVIEWED", "IN_PROGRESS", "CONFIRMED", "REVIEW_REQUIRED")

NAVY = "17365D"
BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
ORANGE = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
RED = "F4CCCC"

PROJECT_SOURCE_HEADERS = [
    "검수순위",
    "classification_project_id",
    "부처코드",
    "부처명",
    "회계유형",
    "프로그램코드",
    "프로그램명",
    "단위사업코드",
    "단위사업명",
    "세부사업코드",
    "세부사업명",
    "관측연도",
    "2022_본예산",
    "2023_본예산",
    "2024_본예산",
    "2025_본예산",
    "4년_본예산합계",
    "UNKNOWN_누적커버리지",
]
PROJECT_INPUT_HEADERS = [
    "analysis_scope_status",
    "scope_exclusion_reason",
    "fiscal_instrument_applicability",
    "fiscal_instrument",
    "all_years_same_classification",
    "classification_evidence",
    "evidence_source",
    "confidence",
    "reviewer",
    "reviewed_at",
    "review_status",
    "review_note",
]
PROJECT_DERIVED_HEADERS = [
    "comparison_group",
    "ranking_population_impact",
    "input_check",
]

YEAR_SOURCE_HEADERS = [
    "classification_project_id",
    "부처명",
    "프로그램명",
    "단위사업명",
    "세부사업명",
    "회계연도",
    "연도별_본예산",
]
YEAR_INPUT_HEADERS = [
    "year_scope_status_override",
    "year_scope_exclusion_reason_override",
    "year_instrument_applicability_override",
    "year_fiscal_instrument_override",
    "year_classification_evidence",
    "year_evidence_source",
    "year_review_status",
    "year_review_note",
]


class UnknownReviewError(ValueError):
    """검수 워크북 입력이나 구조가 유효하지 않을 때 발생합니다."""


@dataclass(frozen=True)
class UnknownReviewPaths:
    priority: Path
    ranking_population: Path
    workbook: Path
    validation_summary: Path

    @classmethod
    def from_root(cls, root: Path) -> UnknownReviewPaths:
        return cls(
            priority=root / "data/analytics/m3/unknown_manual_review_priority.csv",
            ranking_population=root
            / "data/processed/masters/population_sensitivity/ranking_population_v2.parquet",
            workbook=root / "data/manual/unknown_priority_fiscal_instrument_review.xlsx",
            validation_summary=root
            / "data/manual/unknown_priority_fiscal_instrument_review_validation.json",
        )


@dataclass(frozen=True)
class ValidationResult:
    status: str
    project_count: int
    year_row_count: int
    confirmed_project_count: int
    error_count: int
    warning_count: int
    errors: tuple[str, ...]
    warnings: tuple[str, ...]

    def as_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "project_count": self.project_count,
            "year_row_count": self.year_row_count,
            "confirmed_project_count": self.confirmed_project_count,
            "error_count": self.error_count,
            "warning_count": self.warning_count,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
        }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _as_bool(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.lower().isin({"true", "1", "yes"})


def _prepare_source_frames(
    paths: UnknownReviewPaths,
    *,
    expected_project_count: int | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not paths.priority.exists():
        raise UnknownReviewError(f"UNKNOWN 우선검토 파일이 없습니다: {paths.priority}")
    if not paths.ranking_population.exists():
        raise UnknownReviewError(f"ranking population v2가 없습니다: {paths.ranking_population}")

    priority = pd.read_csv(paths.priority, dtype={"ministry_code": "string"})
    required_priority = {
        "classification_project_id",
        "budget_coverage_order",
        "priority_80pct_coverage",
        "cumulative_unknown_budget_share",
    }
    missing_priority = required_priority - set(priority.columns)
    if missing_priority:
        raise UnknownReviewError(
            f"UNKNOWN 우선검토 필수 컬럼이 없습니다: {sorted(missing_priority)}"
        )
    selected = priority[_as_bool(priority["priority_80pct_coverage"])].copy()
    selected = selected.sort_values("budget_coverage_order")
    if expected_project_count is not None and len(selected) != expected_project_count:
        raise UnknownReviewError(
            f"예상 검수 사업 {expected_project_count}개와 실제 {len(selected)}개가 다릅니다."
        )
    if selected["classification_project_id"].duplicated().any():
        raise UnknownReviewError("상위 검수 목록의 classification_project_id가 중복됩니다.")

    population = pd.read_parquet(paths.ranking_population)
    required_population = {
        "classification_project_id",
        "fiscal_year",
        "ministry_code",
        "analysis_ministry_name",
        "account_type_classified",
        "program_code",
        "program_name",
        "activity_code",
        "activity_name",
        "subactivity_code",
        "subactivity_name",
        "original_budget_analysis_amount",
    }
    missing_population = required_population - set(population.columns)
    if missing_population:
        raise UnknownReviewError(
            f"ranking population 필수 컬럼이 없습니다: {sorted(missing_population)}"
        )
    population = population[
        population["classification_project_id"].isin(selected["classification_project_id"])
    ].copy()
    population["fiscal_year"] = pd.to_numeric(population["fiscal_year"], errors="raise").astype(int)
    population = population.sort_values(["classification_project_id", "fiscal_year"])
    if set(population["classification_project_id"]) != set(selected["classification_project_id"]):
        raise UnknownReviewError("상위 검수 목록 일부가 ranking population에서 누락됐습니다.")

    if population.duplicated(["classification_project_id", "fiscal_year"]).any():
        raise UnknownReviewError("사업-연도 행이 중복됩니다.")
    yearly_counts = population.groupby("classification_project_id")["fiscal_year"].nunique()
    if yearly_counts.lt(1).any():
        raise UnknownReviewError("관측연도가 없는 사업이 있습니다.")
    return selected, population


def _set_title(sheet: Any, title: str, subtitle: str, end_column: int) -> None:
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end_letter}1")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(f"A2:{end_letter}2")
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet["A2"].font = Font(color=NAVY, italic=True)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34
    sheet.sheet_view.showGridLines = False


def _style_headers(sheet: Any, headers: list[str], source_count: int, input_count: int) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(HEADER_ROW, column, header)
        if column <= source_count:
            fill = BLUE
        elif column <= source_count + input_count:
            fill = YELLOW
        else:
            fill = GREEN
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=NAVY, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[HEADER_ROW].height = 38
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW}"
    sheet.freeze_panes = f"F{DATA_START_ROW}"


def _add_list_validation(
    sheet: Any,
    column: int,
    start_row: int,
    end_row: int,
    formula: str,
) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "코드값 시트의 허용값 중 하나를 선택해 주세요."
    validation.errorTitle = "허용되지 않는 값"
    validation.prompt = "드롭다운에서 선택해 주세요."
    validation.promptTitle = "입력값 선택"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(column)}{start_row}:{get_column_letter(column)}{end_row}")


def _write_codes_sheet(workbook: Workbook) -> dict[str, str]:
    sheet = workbook.create_sheet(SHEET_CODES)
    _set_title(
        sheet,
        "검수 입력 코드값",
        "검수 시 임의 문구 대신 아래 코드값을 사용합니다. 설명은 판단 기준이며 자동 확정 규칙이 아닙니다.",
        3,
    )
    groups = [
        (
            "analysis_scope_status",
            SCOPE_STATUSES,
            {
                "IN_SCOPE": "정책사업 점검 우선순위 분석 범위에 포함",
                "OUT_OF_SCOPE": "내부거래·금융자산운용·원금상환 등 범위 밖",
                "REVIEW_REQUIRED": "공식 근거가 부족해 추가 확인 필요",
            },
        ),
        (
            "scope_exclusion_reason",
            SCOPE_EXCLUSION_REASONS,
            {
                "INTERNAL_TRANSFER": "회계·기금 간 내부거래 또는 전출",
                "FINANCIAL_ASSET_OPERATION": "국채·주식·예치 등 금융자산 운용",
                "DEBT_REPAYMENT": "차입금·원금 상환",
                "RESERVE_OR_SURPLUS_MANAGEMENT": "여유자금·잉여금 운용",
                "NON_POLICY_ADMINISTRATION": "정책사업 순위와 다른 행정·관리 성격",
                "OUTSIDE_TARGET_POLICY_SCOPE": "현재 분석 질문의 정책사업 범위 밖",
                "OTHER": "기타 사유이며 review_note에 상세 기재",
                "NOT_APPLICABLE": "IN_SCOPE여서 제외 사유가 적용되지 않음",
                "REVIEW_REQUIRED": "제외 사유를 아직 확정할 수 없음",
            },
        ),
        (
            "fiscal_instrument_applicability",
            APPLICABILITY_VALUES,
            {
                "APPLICABLE": "재정수단 분류 가능",
                "NOT_APPLICABLE": "범위 밖 또는 재정수단 분류가 부적절",
                "REVIEW_REQUIRED": "적용 가능 여부 추가 확인 필요",
            },
        ),
        (
            "fiscal_instrument",
            FISCAL_INSTRUMENTS,
            {value: "저장소 재정수단 허용값" for value in FISCAL_INSTRUMENTS},
        ),
        (
            "all_years_same_classification",
            YES_NO_REVIEW,
            {
                "YES": "모든 관측연도에 동일 판정",
                "NO": "연도별 판정 차이가 있어 연도별확인 시트 작성",
                "REVIEW_REQUIRED": "연도별 일관성 추가 확인 필요",
            },
        ),
        (
            "confidence",
            CONFIDENCE_VALUES,
            {
                "HIGH": "공식 문서·페이지에서 사업 성격을 직접 확인",
                "MEDIUM": "복수 공식 자료가 일치하지만 직접 정의가 제한적",
                "LOW": "명칭·간접 자료 중심으로 단독 근거 사용 금지",
            },
        ),
        (
            "review_status",
            REVIEW_STATUSES,
            {
                "UNREVIEWED": "검수 전",
                "IN_PROGRESS": "근거 확인 중",
                "CONFIRMED": "필수 근거와 판단 완료",
                "REVIEW_REQUIRED": "판단 보류",
            },
        ),
    ]
    locations: dict[str, str] = {}
    row = 4
    for name, values, descriptions in groups:
        sheet.cell(row, 1, "입력 필드")
        sheet.cell(row, 2, "허용값")
        sheet.cell(row, 3, "설명")
        for column in range(1, 4):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=BLUE)
            sheet.cell(row, column).font = Font(color=NAVY, bold=True)
        start_row = row + 1
        for value in values:
            row += 1
            sheet.cell(row, 1, name)
            sheet.cell(row, 2, value)
            sheet.cell(row, 3, descriptions.get(value, ""))
        locations[name] = f"'{SHEET_CODES}'!$B${start_row}:$B${row}"
        row += 2
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 64
    for row_number in range(4, row + 1):
        sheet.row_dimensions[row_number].height = 24
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A4"
    return locations


def _write_guide_sheet(workbook: Workbook, project_count: int, year_count: int) -> None:
    sheet = workbook.active
    sheet.title = SHEET_GUIDE
    _set_title(
        sheet,
        "UNKNOWN 예산 80% 커버리지 사람 검수 워크북",
        "노란색 칸만 입력합니다. 범위 판단을 먼저 하고, 범위 안에서만 재정수단을 확정합니다.",
        9,
    )
    sections = [
        ("1. 검수 대상", f"고유 사업 {project_count}개, 사업-연도 {year_count}행"),
        (
            "2. 입력 순서",
            (
                "analysis_scope_status → scope_exclusion_reason → "
                "fiscal_instrument_applicability → fiscal_instrument → "
                "연도 동일성 → 근거·검수상태"
            ),
        ),
        (
            "3. 공식 근거",
            "예산서·기금운용계획·사업설명자료 등 공식 자료명과 페이지 또는 URL을 evidence_source에 기록합니다.",
        ),
        (
            "4. 범위 밖 사업",
            (
                "내부거래·전출·금융자산운용·원금상환이면 재정수단을 "
                "억지로 부여하지 말고 OUT_OF_SCOPE와 제외 사유를 먼저 기록합니다."
            ),
        ),
        (
            "5. 연도 차이",
            (
                "모든 관측연도에 동일하면 YES를 선택합니다. 하나라도 다르면 NO를 "
                "선택하고 연도별확인 시트에서 달라지는 연도를 기록합니다."
            ),
        ),
        (
            "6. 확정 조건",
            "CONFIRMED는 판단값·근거·출처·확신도·검수자·검수일이 모두 있을 때만 선택합니다.",
        ),
        (
            "7. 자동 계산",
            "comparison_group과 ranking_population_impact는 사용자 입력값이 아니며 검수 반영 단계에서 계산합니다.",
        ),
        (
            "8. 금지",
            "명칭만 보고 확정하지 않으며, 애매한 값은 UNKNOWN 추정 대신 REVIEW_REQUIRED로 남깁니다.",
        ),
    ]
    row = 4
    for label, text in sections:
        sheet.cell(row, 1, label)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE)
        sheet.cell(row, 1).font = Font(color=NAVY, bold=True)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        sheet.cell(row, 2, text)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 38
        row += 1

    row += 1
    sheet.cell(row, 1, "색상")
    sheet.cell(row, 2, "의미")
    for cell in sheet[row][0:2]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    legend = [
        (BLUE, "원천값·수정하지 않음"),
        (YELLOW, "사용자 입력"),
        (GREEN, "후속 코드 자동계산"),
        (ORANGE, "추가 검토 필요"),
    ]
    for fill, meaning in legend:
        row += 1
        sheet.cell(row, 1, "")
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=fill)
        sheet.cell(row, 2, meaning)
    sheet.column_dimensions["A"].width = 26
    for column in range(2, 10):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    sheet.freeze_panes = "A4"


def _project_rows(selected: pd.DataFrame, population: pd.DataFrame) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for priority in selected.itertuples(index=False):
        project_id = str(priority.classification_project_id)
        part = population[population["classification_project_id"].eq(project_id)].copy()
        latest = part.sort_values("fiscal_year").iloc[-1]
        budgets = (
            part.groupby("fiscal_year")["original_budget_analysis_amount"]
            .sum(min_count=1)
            .to_dict()
        )
        rows.append(
            [
                int(priority.budget_coverage_order),
                project_id,
                str(latest["ministry_code"]),
                latest["analysis_ministry_name"],
                latest["account_type_classified"],
                str(latest["program_code"]),
                latest["program_name"],
                str(latest["activity_code"]),
                latest["activity_name"],
                str(latest["subactivity_code"]),
                latest["subactivity_name"],
                ";".join(map(str, sorted(part["fiscal_year"].astype(int).unique()))),
                budgets.get(2022),
                budgets.get(2023),
                budgets.get(2024),
                budgets.get(2025),
                pd.to_numeric(part["original_budget_analysis_amount"], errors="coerce").sum(
                    min_count=1
                ),
                float(priority.cumulative_unknown_budget_share),
            ]
        )
    return rows


def _write_project_sheet(
    workbook: Workbook,
    selected: pd.DataFrame,
    population: pd.DataFrame,
    validation_locations: dict[str, str],
) -> None:
    sheet = workbook.create_sheet(SHEET_PROJECTS)
    headers = PROJECT_SOURCE_HEADERS + PROJECT_INPUT_HEADERS + PROJECT_DERIVED_HEADERS
    _set_title(
        sheet,
        f"사업 기본판정 — UNKNOWN 예산 80% 커버리지 {len(selected)}개",
        "파란색은 원천값, 노란색은 사용자 입력, 초록색은 후속 자동계산입니다. "
        "범위 판정 후 재정수단을 검수해 주세요.",
        len(headers),
    )
    _style_headers(
        sheet,
        headers,
        source_count=len(PROJECT_SOURCE_HEADERS),
        input_count=len(PROJECT_INPUT_HEADERS),
    )
    rows = _project_rows(selected, population)
    for row_number, values in enumerate(rows, start=DATA_START_ROW):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, None if pd.isna(value) else value)
            sheet.cell(row_number, column).fill = PatternFill("solid", fgColor=PALE_BLUE)
        for column in range(
            len(PROJECT_SOURCE_HEADERS) + 1,
            len(PROJECT_SOURCE_HEADERS) + len(PROJECT_INPUT_HEADERS) + 1,
        ):
            sheet.cell(row_number, column).fill = PatternFill("solid", fgColor=YELLOW)
        for column in range(
            len(PROJECT_SOURCE_HEADERS) + len(PROJECT_INPUT_HEADERS) + 1,
            len(headers) + 1,
        ):
            sheet.cell(row_number, column).fill = PatternFill("solid", fgColor=GREEN)
        sheet.cell(row_number, headers.index("comparison_group") + 1, "검수 후 자동계산")
        sheet.cell(row_number, headers.index("ranking_population_impact") + 1, "검수 후 자동계산")

        def col(name: str) -> str:
            return get_column_letter(headers.index(name) + 1)

        check_formula = (
            f'=IF({col("review_status")}{row_number}="","미입력",'
            f'IF({col("review_status")}{row_number}="UNREVIEWED","미입력",'
            f'IF(AND({col("review_status")}{row_number}="CONFIRMED",'
            f'OR({col("analysis_scope_status")}{row_number}="",'
            f'{col("fiscal_instrument_applicability")}{row_number}="",'
            f'{col("all_years_same_classification")}{row_number}="",'
            f'{col("classification_evidence")}{row_number}="",'
            f'{col("evidence_source")}{row_number}="",'
            f'{col("confidence")}{row_number}="",'
            f'{col("reviewer")}{row_number}="",'
            f'{col("reviewed_at")}{row_number}="")),"필수값 누락",'
            f'IF(AND({col("analysis_scope_status")}{row_number}="OUT_OF_SCOPE",'
            f'{col("scope_exclusion_reason")}{row_number}=""),"제외사유 필요",'
            f'IF(AND({col("analysis_scope_status")}{row_number}="IN_SCOPE",'
            f'{col("fiscal_instrument_applicability")}{row_number}="APPLICABLE",'
            f'OR({col("fiscal_instrument")}{row_number}="",'
            f'{col("fiscal_instrument")}{row_number}="UNKNOWN")),'
            f'"재정수단 확정 필요","확인")))))'
        )
        sheet.cell(row_number, headers.index("input_check") + 1, check_formula)

    end_row = DATA_START_ROW + len(rows) - 1
    validation_map = {
        "analysis_scope_status": "analysis_scope_status",
        "scope_exclusion_reason": "scope_exclusion_reason",
        "fiscal_instrument_applicability": "fiscal_instrument_applicability",
        "fiscal_instrument": "fiscal_instrument",
        "all_years_same_classification": "all_years_same_classification",
        "confidence": "confidence",
        "review_status": "review_status",
    }
    for header, code_group in validation_map.items():
        _add_list_validation(
            sheet,
            headers.index(header) + 1,
            DATA_START_ROW,
            end_row,
            validation_locations[code_group],
        )

    status_column = get_column_letter(headers.index("review_status") + 1)
    whole_range = f"A{DATA_START_ROW}:{get_column_letter(len(headers))}{end_row}"
    sheet.conditional_formatting.add(
        whole_range,
        FormulaRule(
            formula=[f'${status_column}{DATA_START_ROW}="CONFIRMED"'],
            fill=PatternFill("solid", fgColor=GREEN),
        ),
    )
    sheet.conditional_formatting.add(
        whole_range,
        FormulaRule(
            formula=[f'${status_column}{DATA_START_ROW}="REVIEW_REQUIRED"'],
            fill=PatternFill("solid", fgColor=ORANGE),
        ),
    )
    for column in range(1, len(headers) + 1):
        letter = get_column_letter(column)
        width = 15
        if headers[column - 1] in {
            "classification_project_id",
            "classification_evidence",
            "evidence_source",
            "review_note",
        }:
            width = 34
        elif "예산" in headers[column - 1]:
            width = 24
        elif headers[column - 1].endswith("명"):
            width = 22
        sheet.column_dimensions[letter].width = width
    for row_number in range(DATA_START_ROW, end_row + 1):
        sheet.row_dimensions[row_number].height = 42
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for header in ["2022_본예산", "2023_본예산", "2024_본예산", "2025_본예산", "4년_본예산합계"]:
        column = get_column_letter(headers.index(header) + 1)
        for cell in sheet[f"{column}{DATA_START_ROW}:{column}{end_row}"]:
            cell[0].number_format = '#,##0" 원"'
    share_column = get_column_letter(headers.index("UNKNOWN_누적커버리지") + 1)
    for cell in sheet[f"{share_column}{DATA_START_ROW}:{share_column}{end_row}"]:
        cell[0].number_format = "0.0%"
    date_column = get_column_letter(headers.index("reviewed_at") + 1)
    for cell in sheet[f"{date_column}{DATA_START_ROW}:{date_column}{end_row}"]:
        cell[0].number_format = "yyyy-mm-dd"


def _write_year_sheet(
    workbook: Workbook,
    population: pd.DataFrame,
    validation_locations: dict[str, str],
) -> None:
    sheet = workbook.create_sheet(SHEET_YEARS)
    headers = YEAR_SOURCE_HEADERS + YEAR_INPUT_HEADERS
    _set_title(
        sheet,
        (
            "연도별 적용 확인 — "
            f"{population['classification_project_id'].nunique()}개 사업, "
            f"실제 관측 {len(population)}행"
        ),
        "사업검수의 all_years_same_classification이 NO일 때만 해당 사업의 연도별 노란색 칸을 작성합니다.",
        len(headers),
    )
    _style_headers(
        sheet,
        headers,
        source_count=len(YEAR_SOURCE_HEADERS),
        input_count=len(YEAR_INPUT_HEADERS),
    )
    population = population.sort_values(["classification_project_id", "fiscal_year"])
    for row_number, row in enumerate(population.itertuples(index=False), start=DATA_START_ROW):
        values = [
            row.classification_project_id,
            row.analysis_ministry_name,
            row.program_name,
            row.activity_name,
            row.subactivity_name,
            int(row.fiscal_year),
            row.original_budget_analysis_amount,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, None if pd.isna(value) else value)
            sheet.cell(row_number, column).fill = PatternFill("solid", fgColor=PALE_BLUE)
        for column in range(len(YEAR_SOURCE_HEADERS) + 1, len(headers) + 1):
            sheet.cell(row_number, column).fill = PatternFill("solid", fgColor=YELLOW)

    end_row = DATA_START_ROW + len(population) - 1
    validation_map = {
        "year_scope_status_override": "analysis_scope_status",
        "year_scope_exclusion_reason_override": "scope_exclusion_reason",
        "year_instrument_applicability_override": "fiscal_instrument_applicability",
        "year_fiscal_instrument_override": "fiscal_instrument",
        "year_review_status": "review_status",
    }
    for header, code_group in validation_map.items():
        _add_list_validation(
            sheet,
            headers.index(header) + 1,
            DATA_START_ROW,
            end_row,
            validation_locations[code_group],
        )
    widths = [38, 18, 22, 24, 28, 11, 24, 24, 28, 28, 24, 38, 38, 20, 38]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(DATA_START_ROW, end_row + 1):
        sheet.row_dimensions[row_number].height = 36
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[f"G{DATA_START_ROW}:G{end_row}"]:
        cell[0].number_format = '#,##0" 원"'


def _write_meta_sheet(
    workbook: Workbook,
    paths: UnknownReviewPaths,
    selected: pd.DataFrame,
    population: pd.DataFrame,
) -> None:
    sheet = workbook.create_sheet(SHEET_META)
    _set_title(
        sheet,
        "검수 원본 메타데이터",
        "검수 파일이 어떤 입력과 기준으로 생성됐는지 추적합니다. 이 시트는 수정하지 않습니다.",
        4,
    )
    rows = [
        ("schema_version", 1),
        ("generated_at_utc", datetime.now(UTC).isoformat()),
        ("selection_rule", "priority_80pct_coverage=true"),
        ("project_count", len(selected)),
        ("project_year_row_count", len(population)),
        ("priority_source", str(paths.priority)),
        ("priority_source_sha256", _sha256(paths.priority)),
        ("ranking_population_source", str(paths.ranking_population)),
        ("ranking_population_sha256", _sha256(paths.ranking_population)),
        ("amount_field", "original_budget_analysis_amount"),
        ("analysis_unit", "classification_project_id 및 classification_project_id×fiscal_year"),
        ("warning", "이 파일은 정책판정이 아니라 점검 후보 분석을 위한 사람 검수 입력입니다."),
    ]
    sheet["A4"] = "항목"
    sheet["B4"] = "값"
    for cell in sheet[4][0:2]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=NAVY, bold=True)
    for row_number, (key, value) in enumerate(rows, start=5):
        sheet.cell(row_number, 1, key)
        sheet.cell(row_number, 2, value)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 110
    for row_number in range(5, 5 + len(rows)):
        sheet.cell(row_number, 2).alignment = Alignment(
            horizontal="left", wrap_text=True, vertical="top"
        )


def build_unknown_review_workbook(
    paths: UnknownReviewPaths,
    *,
    overwrite: bool = False,
    expected_project_count: int | None = None,
) -> Path:
    """상위 UNKNOWN 사업 검수용 보존형 Excel 워크북을 생성합니다."""
    if paths.workbook.exists() and not overwrite:
        raise UnknownReviewError(
            f"검수 워크북이 이미 있습니다. 사람 입력 보호를 위해 덮어쓰지 않습니다: {paths.workbook}"
        )
    selected, population = _prepare_source_frames(
        paths, expected_project_count=expected_project_count
    )
    workbook = Workbook()
    _write_guide_sheet(workbook, len(selected), len(population))
    validation_locations = _write_codes_sheet(workbook)
    _write_project_sheet(workbook, selected, population, validation_locations)
    _write_year_sheet(workbook, population, validation_locations)
    _write_meta_sheet(workbook, paths, selected, population)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    paths.workbook.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(paths.workbook)
    return paths.workbook


def _records_from_sheet(sheet: Any) -> list[dict[str, Any]]:
    headers = [cell.value for cell in sheet[HEADER_ROW]]
    records = []
    for values in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not any(value is not None for value in values):
            continue
        records.append(dict(zip(headers, values, strict=True)))
    return records


def apply_unknown_review_overlay(
    frame: pd.DataFrame,
    workbook_path: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """완료된 사업 단위 검수를 분석 프레임에 적용하고 변경 내역을 반환합니다."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    decisions = pd.DataFrame(_records_from_sheet(workbook[SHEET_PROJECTS]))
    if decisions.empty or not decisions["review_status"].eq("CONFIRMED").all():
        raise UnknownReviewError("오버레이에는 모든 사업의 CONFIRMED 검수가 필요합니다.")
    if not decisions["all_years_same_classification"].eq("YES").all():
        raise UnknownReviewError(
            "연도별 분류가 다른 사업은 현재 오버레이 대상이 아닙니다. 연도별 override를 먼저 확정하세요."
        )
    if decisions["classification_project_id"].duplicated().any():
        raise UnknownReviewError("오버레이 사업 ID가 중복됩니다.")

    result = frame.copy()
    review_ids = set(decisions["classification_project_id"].astype(str))
    observed_ids = set(result["classification_project_id"].astype(str))
    missing = review_ids - observed_ids
    if missing:
        raise UnknownReviewError(f"분석 프레임에 없는 검수 사업이 있습니다: {sorted(missing)[:3]}")

    result["unknown_review_applied"] = False
    result["unknown_review_status"] = pd.NA
    result["unknown_review_evidence_source"] = pd.NA
    audit_rows: list[dict[str, Any]] = []
    for _, decision in decisions.iterrows():
        project_id = str(decision["classification_project_id"])
        mask = result["classification_project_id"].astype(str).eq(project_id)
        scope = str(decision["analysis_scope_status"])
        applicability = str(decision["fiscal_instrument_applicability"])
        instrument = (
            str(decision["fiscal_instrument"]) if applicability == "APPLICABLE" else "UNKNOWN"
        )
        before = result.loc[mask]
        result.loc[mask, "analysis_included_classified"] = scope == "IN_SCOPE"
        result.loc[mask, "exclusion_category_classified"] = (
            pd.NA if scope == "IN_SCOPE" else decision["scope_exclusion_reason"]
        )
        result.loc[mask, "exclusion_reason_classified"] = (
            pd.NA if scope == "IN_SCOPE" else decision["classification_evidence"]
        )
        result.loc[mask, "fiscal_instrument"] = instrument
        result.loc[mask, "instrument_classification_method"] = (
            "MANUAL_OFFICIAL_EVIDENCE"
            if applicability == "APPLICABLE"
            else "MANUAL_STRUCTURAL_MIXED_OR_NOT_APPLICABLE"
        )
        result.loc[mask, "instrument_classification_evidence"] = decision["classification_evidence"]
        result.loc[mask, "instrument_manual_review_required"] = False
        result.loc[mask, "classification_manual_review_required"] = False
        result.loc[mask, "classification_status"] = "MANUAL_CONFIRMED"
        result.loc[mask, "classification_method"] = "MANUAL_OFFICIAL_EVIDENCE"
        result.loc[mask, "classification_evidence"] = decision["classification_evidence"]
        result.loc[mask, "comparison_group"] = (
            result.loc[mask, "account_type_classified"].astype(str)
            + "|"
            + instrument
            + "|"
            + result.loc[mask, "project_category"].astype(str)
        )
        result.loc[mask, "unknown_review_applied"] = True
        result.loc[mask, "unknown_review_status"] = "CONFIRMED"
        result.loc[mask, "unknown_review_evidence_source"] = decision["evidence_source"]
        audit_rows.append(
            {
                "classification_project_id": project_id,
                "project_year_row_count": int(mask.sum()),
                "analysis_scope_status": scope,
                "fiscal_instrument_applicability": applicability,
                "fiscal_instrument_before": ";".join(
                    sorted(before["fiscal_instrument"].dropna().astype(str).unique())
                ),
                "fiscal_instrument_after": instrument,
                "ranking_population_impact": (
                    "EXCLUDE_OVERALL_RANKING"
                    if scope == "OUT_OF_SCOPE"
                    else (
                        "KEEP_GENERAL_EXCLUDE_INSTRUMENT_PEER"
                        if instrument == "UNKNOWN"
                        else "KEEP_AND_ASSIGN_INSTRUMENT_PEER"
                    )
                ),
                "original_budget_amount": int(
                    pd.to_numeric(before["original_budget_analysis_amount"], errors="coerce").sum()
                ),
                "evidence_source": decision["evidence_source"],
            }
        )

    eligible = result["analysis_included_classified"].astype("boolean").fillna(False)
    group_sizes = (
        result.loc[eligible].groupby("comparison_group")["classification_project_id"].nunique()
    )
    result["comparison_group_size"] = (
        result["comparison_group"].map(group_sizes).fillna(0).astype("Int64")
    )
    result["ranking_small_group_limited_flag"] = result["comparison_group_size"].lt(5)
    return result, pd.DataFrame(audit_rows)


def validate_unknown_review_workbook(
    paths: UnknownReviewPaths,
    *,
    require_complete: bool = False,
    expected_project_count: int | None = None,
) -> ValidationResult:
    """검수 워크북의 구조, 허용값, 확정행 필수 근거를 검증합니다."""
    if not paths.workbook.exists():
        raise UnknownReviewError(f"검수 워크북이 없습니다: {paths.workbook}")
    selected, current_population = _prepare_source_frames(
        paths, expected_project_count=expected_project_count
    )
    expected_project_count = len(selected)
    expected_year_row_count = len(current_population)
    workbook = load_workbook(paths.workbook, data_only=False)
    required_sheets = {SHEET_GUIDE, SHEET_PROJECTS, SHEET_YEARS, SHEET_CODES, SHEET_META}
    missing_sheets = required_sheets - set(workbook.sheetnames)
    if missing_sheets:
        raise UnknownReviewError(f"필수 시트가 없습니다: {sorted(missing_sheets)}")
    projects = _records_from_sheet(workbook[SHEET_PROJECTS])
    years = _records_from_sheet(workbook[SHEET_YEARS])
    errors: list[str] = []
    warnings: list[str] = []

    if len(projects) != expected_project_count:
        errors.append(f"사업 행 수가 {expected_project_count}개가 아닙니다: {len(projects)}")
    project_ids = [str(row.get("classification_project_id") or "") for row in projects]
    if len(set(project_ids)) != len(project_ids):
        errors.append("classification_project_id가 중복됩니다.")
    expected_project_ids = set(selected["classification_project_id"].astype(str))
    if set(project_ids) != expected_project_ids:
        errors.append("현재 UNKNOWN 우선검토 목록과 워크북 사업 목록이 다릅니다.")
    if len(years) != expected_year_row_count:
        errors.append(
            f"사업-연도 행 수가 현재 관측 {expected_year_row_count}개가 아닙니다: {len(years)}"
        )

    allowed = {
        "analysis_scope_status": set(SCOPE_STATUSES),
        "scope_exclusion_reason": set(SCOPE_EXCLUSION_REASONS),
        "fiscal_instrument_applicability": set(APPLICABILITY_VALUES),
        "fiscal_instrument": set(FISCAL_INSTRUMENTS),
        "all_years_same_classification": set(YES_NO_REVIEW),
        "confidence": set(CONFIDENCE_VALUES),
        "review_status": set(REVIEW_STATUSES),
    }
    confirmed_count = 0
    for row in projects:
        project_id = str(row.get("classification_project_id") or "")
        for field, values in allowed.items():
            value = row.get(field)
            if value not in (None, "") and str(value) not in values:
                errors.append(f"{project_id}: {field} 허용값 위반: {value}")
        status = row.get("review_status")
        if status == "CONFIRMED":
            confirmed_count += 1
            required = [
                "analysis_scope_status",
                "fiscal_instrument_applicability",
                "all_years_same_classification",
                "classification_evidence",
                "evidence_source",
                "confidence",
                "reviewer",
                "reviewed_at",
            ]
            missing = [field for field in required if row.get(field) in (None, "")]
            if missing:
                errors.append(f"{project_id}: CONFIRMED 필수값 누락: {', '.join(missing)}")
            scope = row.get("analysis_scope_status")
            applicability = row.get("fiscal_instrument_applicability")
            instrument = row.get("fiscal_instrument")
            if scope == "OUT_OF_SCOPE" and row.get("scope_exclusion_reason") in (None, ""):
                errors.append(f"{project_id}: OUT_OF_SCOPE 제외 사유가 없습니다.")
            if (
                scope == "IN_SCOPE"
                and applicability == "APPLICABLE"
                and instrument in (None, "", "UNKNOWN")
            ):
                errors.append(f"{project_id}: IN_SCOPE·APPLICABLE 재정수단이 미확정입니다.")
            if applicability == "NOT_APPLICABLE" and instrument not in (None, "", "UNKNOWN"):
                errors.append(
                    f"{project_id}: NOT_APPLICABLE인데 재정수단 {instrument}이 입력됐습니다."
                )
            if row.get("all_years_same_classification") == "NO":
                project_years = [
                    year
                    for year in years
                    if str(year.get("classification_project_id") or "") == project_id
                ]
                confirmed_years = [
                    year for year in project_years if year.get("year_review_status") == "CONFIRMED"
                ]
                has_override = any(
                    any(
                        year.get(field) not in (None, "")
                        for field in [
                            "year_scope_status_override",
                            "year_scope_exclusion_reason_override",
                            "year_instrument_applicability_override",
                            "year_fiscal_instrument_override",
                        ]
                    )
                    for year in project_years
                )
                if len(confirmed_years) != len(project_years):
                    errors.append(
                        f"{project_id}: 연도별 분류가 다르지만 모든 관측연도 확인이 완료되지 않았습니다."
                    )
                if not has_override:
                    errors.append(
                        f"{project_id}: all_years_same_classification=NO인데 연도별 변경값이 없습니다."
                    )
        elif status in (None, "", "UNREVIEWED"):
            warnings.append(f"{project_id}: 아직 검수되지 않았습니다.")
        elif status == "IN_PROGRESS":
            warnings.append(f"{project_id}: 검수가 진행 중입니다.")
        elif status == "REVIEW_REQUIRED":
            warnings.append(f"{project_id}: 추가 검토가 필요합니다.")

    if require_complete and confirmed_count != expected_project_count:
        errors.append(
            f"완료 검증에는 {expected_project_count}개 CONFIRMED가 필요하지만 "
            f"{confirmed_count}개입니다."
        )
    status = "PASS" if not errors and not warnings else "INCOMPLETE"
    if errors:
        status = "FAIL"
    result = ValidationResult(
        status=status,
        project_count=len(projects),
        year_row_count=len(years),
        confirmed_project_count=confirmed_count,
        error_count=len(errors),
        warning_count=len(warnings),
        errors=tuple(errors),
        warnings=tuple(warnings),
    )
    paths.validation_summary.parent.mkdir(parents=True, exist_ok=True)
    paths.validation_summary.write_text(
        json.dumps(result.as_dict(), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return result
