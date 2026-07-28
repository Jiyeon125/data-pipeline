"""중기부 성과계획서·성과보고서 PDF 별첨과 수기 성과지표 63행을 원문 대조합니다.

이 모듈은 2022~2024년 계획서·보고서 별첨과, 같은 파일명의 보고서 전체본이
있으면 본문의 성과지표 상세표를 함께 사용합니다.

- 계획서 별첨1 "프로그램 성과지표 현황": 2022년은 표 셀이 이미지로 렌더링되어
  텍스트 레이어가 없어 OCR이 필요합니다. 2023·2024년은 실제 텍스트가 있습니다.
  (분리 PDF 자체가 별첨1로 시작하도록 잘려 있어 "별첨1" 글자 자체는 연도마다
  있거나 없습니다.)
- 보고서 별첨3 "성과 달성도 현황"의 "3. 세부현황" 표와 보고서 본문:
  부처에 따라 별첨3은 지표명만 있고 목표·실적·달성률 상세표는 본문에 있으므로
  전체본이 있으면 본문의 완전한 표를 우선합니다. 프로그램목표별
  성과지표마다 최근 3개년(예: '22/'23/'24) 목표·실적·달성률을 나열합니다.
  "목표"/"실적"/"달성률" 줄 라벨을 앵커로 삼아 마지막(해당 연도) 값을
  추출합니다. 이 표가 63행과 실제로 대응하는 핵심 대조 표입니다.
  (별첨7 "보조사업에 대한 성과달성 현황"은 보조금 사업 세부지표 중심이라
  63행의 프로그램 목표 지표 다수를 포함하지 않아 사용하지 않습니다.)
- 보고서 별첨6 "OO년도 성과계획서 변경 사항": 정부안과 확정예산 사이의
  프로그램목표별 지표·목표치·예산 변경 내역이 실제 텍스트로 존재합니다.
  계획서 별첨1의 확정(변경 후) 값을 실제 텍스트로 교차 검증하고, 별첨1이
  OCR 대상인 2022년에는 텍스트 기반 보강 근거로도 사용합니다.

원본 PDF와 수기 엑셀은 읽기 전용으로만 열고 수정하지 않습니다.
외부 LLM API는 호출하지 않습니다. 로컬 텍스트 추출과 로컬 Tesseract OCR만
사용합니다.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

APPENDIX_ROOT = Path("data/raw/performance_docs/appendix")
PERFORMANCE_DOC_ROOT = APPENDIX_ROOT.parent
LOW_TEXT_CHAR_THRESHOLD = 100
TESSERACT_LANG = "kor+eng"
TESSERACT_PSM = "6"

# 로컬 OCR 실행 환경 (관리자 권한 없이 설치한 사용자 tessdata 우선 사용).
DEFAULT_TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
DEFAULT_TESSDATA_DIR = str(Path.home() / ".tessdata")

STANDALONE_NUMERIC_RE = re.compile(r"^[()\-\d,.%]+$")
NUMBER_TOKEN_RE = re.compile(r"\(?-?\d[\d,]*\.?\d*\)?")
WEIGHT_TOKEN_RE = re.compile(r"^[01](?:\.\d)?$")
PRINTED_PAGE_RE = re.compile(r"-\s*(\d{1,4})\s*-")
CIRCLED_DIGIT_RE = re.compile(r"[\u2460-\u2473]")
MISSING_VALUE_TOKENS = {"", "-", "신규", "종료", "집계중", "해당없음", "잠정"}

ALLOWED_STATUS_VALUES = {
    "EXACT_MATCH",
    "MATCH_AFTER_CHANGE",
    "ROUNDING_ONLY",
    "VALUE_MISMATCH",
    "MANUAL_MISSING_PDF_PRESENT",
    "PDF_MISSING_MANUAL_PRESENT",
    "PDF_NOT_FOUND",
    "OCR_REQUIRED",
    "AMBIGUOUS",
    "MANUAL_REVIEW",
    "NOT_APPLICABLE",
}
# 우선순위가 높은 값이 앞에 오며, overall_reconciliation_status 결정에 사용합니다.
# 프롬프트 8절의 7단계(AMBIGUOUS~EXACT_MATCH)를 그대로 지키고, 프롬프트에 없는
# MANUAL_REVIEW·PDF_NOT_FOUND·ROUNDING_ONLY·NOT_APPLICABLE은 다음 근거로
# 배치합니다: MANUAL_REVIEW는 AMBIGUOUS와 동급으로 사람 확인이 필요해 최상위,
# PDF_NOT_FOUND는 MANUAL_MISSING_PDF_PRESENT 다음(둘 다 근거 부재), ROUNDING_ONLY는
# MATCH_AFTER_CHANGE보다 낮은 사소한 수치차이, NOT_APPLICABLE은 판단 대상이
# 없다는 뜻이라 가장 낮은 우선순위입니다.
STATUS_PRIORITY = [
    "AMBIGUOUS",
    "MANUAL_REVIEW",
    "OCR_REQUIRED",
    "VALUE_MISMATCH",
    "MANUAL_MISSING_PDF_PRESENT",
    "PDF_MISSING_MANUAL_PRESENT",
    "PDF_NOT_FOUND",
    "MATCH_AFTER_CHANGE",
    "ROUNDING_ONLY",
    "EXACT_MATCH",
    "NOT_APPLICABLE",
]


class PdfReconciliationError(ValueError):
    """PDF 대조 입력·구조 계약을 위반할 때 발생합니다."""


@dataclass(frozen=True)
class PdfDocSpec:
    """분리 PDF와 원본 PDF 페이지를 잇는 메타데이터."""

    fiscal_year: int
    doc_type: str  # "plan" | "report"
    filename: str
    source_page_start: int
    source_page_end: int
    ministry_code: str = "102"

    @property
    def path(self) -> Path:
        return (
            APPENDIX_ROOT
            / f"year={self.fiscal_year}"
            / f"ministry_code={self.ministry_code}"
            / self.filename
        )

    def source_pdf_page(self, split_pdf_page: int) -> int:
        return self.source_page_start + split_pdf_page - 1


PDF_DOC_SPECS: tuple[PdfDocSpec, ...] = (
    PdfDocSpec(2022, "plan", "2022년도 성과계획서_중소벤처기업부-176-216.pdf", 176, 216),
    PdfDocSpec(2022, "report", "2022년도 성과보고서_중소벤처기업부-170-212.pdf", 170, 212),
    PdfDocSpec(2023, "plan", "2023년도 성과계획서_중소벤처기업부-184-224.pdf", 184, 224),
    PdfDocSpec(2023, "report", "2023년도 성과보고서_중소벤처기업부-179-226.pdf", 179, 226),
    PdfDocSpec(2024, "plan", "2024년도 성과계획서_중소벤처기업부-188-226.pdf", 188, 226),
    PdfDocSpec(2024, "report", "2024년도 성과보고서_중소벤처기업부-159-208.pdf", 159, 208),
    # 2025년도 성과계획서는 파일 인벤토리(페이지 QA)에는 포함하지만, 대응하는
    # 성과보고서가 아직 없어 63행 대조 모집단(63행)에는 사용하지 않습니다.
    PdfDocSpec(2025, "plan", "2025년도 성과계획서_중소벤처기업부-190-231.pdf", 190, 231),
)

# 2022년 계획서 별첨1(세부현황) 표는 셀이 이미지로 렌더링되어 텍스트 레이어가
# 없습니다. 실제 렌더링·텍스트 점검으로 확인한 사실이며 임의 추정이 아닙니다.
PLAN_TABLE_IS_IMAGE_ONLY: dict[int, bool] = {2022: True, 2023: False, 2024: False}


def doc_spec(
    fiscal_year: int,
    doc_type: str,
    doc_specs: tuple[PdfDocSpec, ...] = PDF_DOC_SPECS,
) -> PdfDocSpec:
    for spec in doc_specs:
        if spec.fiscal_year == fiscal_year and spec.doc_type == doc_type:
            return spec
    raise PdfReconciliationError(f"등록되지 않은 문서입니다: {fiscal_year} {doc_type}")


def discover_pdf_doc_specs(
    ministry_code: str,
    *,
    years: tuple[int, ...] = (2022, 2023, 2024),
) -> tuple[PdfDocSpec, ...]:
    """부처별 별첨 폴더에서 연도별 계획서·보고서 파일을 하나씩 찾습니다."""
    ministry_code = str(ministry_code).zfill(3)
    specs: list[PdfDocSpec] = []
    for year in years:
        root = APPENDIX_ROOT / f"year={year}" / f"ministry_code={ministry_code}"
        for doc_type, token in (("plan", "성과계획서"), ("report", "성과보고서")):
            matches = sorted(root.glob(f"*{token}*.pdf"))
            if len(matches) != 1:
                raise PdfReconciliationError(
                    f"{ministry_code} {year} {doc_type} PDF가 1개여야 합니다 "
                    f"(실제 {len(matches)}개): {root}"
                )
            page_match = re.search(r"-(\d+)-(\d+)\.pdf$", matches[0].name)
            if page_match is None:
                raise PdfReconciliationError(
                    f"파일명 끝에서 원본 페이지 범위를 읽을 수 없습니다: {matches[0].name}"
                )
            start, end = map(int, page_match.groups())
            specs.append(
                PdfDocSpec(
                    year,
                    doc_type,
                    matches[0].name,
                    start,
                    end,
                    ministry_code=ministry_code,
                )
            )
    return tuple(specs)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def all_source_hashes(
    manual_excel_path: Path,
    doc_specs: tuple[PdfDocSpec, ...] = PDF_DOC_SPECS,
) -> dict[str, str]:
    """원본 PDF 6개와 수기 엑셀의 SHA-256을 실행 전후 비교하기 위해 계산합니다."""
    hashes = {str(manual_excel_path): sha256_file(manual_excel_path)}
    for spec in doc_specs:
        hashes[str(spec.path)] = sha256_file(spec.path)
        full_path = full_document_path(spec) if spec.doc_type == "report" else None
        if full_path is not None:
            hashes[str(full_path)] = sha256_file(full_path)
    return hashes


# ---------------------------------------------------------------------------
# 텍스트/OCR 추출
# ---------------------------------------------------------------------------


def load_page_texts(path: Path) -> list[str]:
    if not path.is_file():
        raise PdfReconciliationError(f"PDF 파일을 찾을 수 없습니다: {path}")
    with fitz.open(path) as document:
        return [document[i].get_text() for i in range(document.page_count)]


def printed_page_number(page_text: str) -> int | None:
    match = PRINTED_PAGE_RE.search(page_text)
    if not match:
        return None
    return int(match.group(1))


def _configure_pytesseract() -> Any:
    import os

    import pytesseract

    tesseract_cmd = Path(DEFAULT_TESSERACT_CMD)
    if tesseract_cmd.is_file():
        pytesseract.pytesseract.tesseract_cmd = str(tesseract_cmd)
    tessdata_dir = Path(DEFAULT_TESSDATA_DIR)
    if tessdata_dir.is_dir():
        # `--tessdata-dir` CLI 인자는 pytesseract의 Windows 인용부호 처리에서
        # 깨지므로, 환경변수 TESSDATA_PREFIX로 대신 지정합니다.
        os.environ["TESSDATA_PREFIX"] = str(tessdata_dir)
    return pytesseract


def ocr_page_text(path: Path, page_index: int, *, dpi: int = 300) -> str:
    """지정한 페이지를 렌더링해 로컬 Tesseract(kor+eng)로 OCR합니다."""
    from PIL import Image

    pytesseract = _configure_pytesseract()
    with fitz.open(path) as document:
        page = document[page_index]
        pixmap = page.get_pixmap(dpi=dpi)
        image_bytes = pixmap.tobytes("png")
    image = Image.open(io.BytesIO(image_bytes))
    config = f"--psm {TESSERACT_PSM}"
    return pytesseract.image_to_string(image, lang=TESSERACT_LANG, config=config)


# ---------------------------------------------------------------------------
# 페이지 인벤토리 (검증 항목 1·2·6·7·9 대응)
# ---------------------------------------------------------------------------


def build_page_inventory(
    *,
    run_ocr: bool = True,
    doc_specs: tuple[PdfDocSpec, ...] = PDF_DOC_SPECS,
) -> pd.DataFrame:
    """6개 PDF 전체 페이지의 분리·원본·인쇄 페이지, 텍스트량, 추출 방식을 기록합니다."""
    rows: list[dict[str, Any]] = []
    for spec in doc_specs:
        pages = load_page_texts(spec.path)
        for split_pdf_page, text in enumerate(pages, start=1):
            char_count = len(text.strip())
            low_text = char_count < LOW_TEXT_CHAR_THRESHOLD
            extraction_method = "TEXT"
            ocr_text: str | None = None
            if low_text and run_ocr:
                try:
                    ocr_text = ocr_page_text(spec.path, split_pdf_page - 1)
                    extraction_method = "OCR"
                except Exception as exc:  # noqa: BLE001 - OCR 환경 문제를 기록만 함
                    ocr_text = None
                    extraction_method = f"OCR_FAILED:{exc.__class__.__name__}"
            rows.append(
                {
                    "fiscal_year": spec.fiscal_year,
                    "doc_type": spec.doc_type,
                    "file_name": spec.filename,
                    "split_pdf_page": split_pdf_page,
                    "source_pdf_page": spec.source_pdf_page(split_pdf_page),
                    "printed_page": printed_page_number(text),
                    "char_count": char_count,
                    "low_text_flag": low_text,
                    "extraction_method": extraction_method,
                    "ocr_char_count": len(ocr_text.strip()) if ocr_text else None,
                }
            )
    return pd.DataFrame(rows).convert_dtypes()


# ---------------------------------------------------------------------------
# 문자열/숫자 정규화
# ---------------------------------------------------------------------------


def normalize_indicator_name(raw: Any) -> str:
    """공백·줄바꿈·괄호·순번 원문자·일반 문장부호만 제거한 비교용 정규화명."""
    if raw is None:
        return ""
    text = str(raw)
    # NFKC 정규화는 원문자 숫자(①→1)를 먼저 분해해 버려 아래 제거 정규식이
    # 무력화됩니다. 분해되기 전에 원문자 숫자를 먼저 제거합니다.
    text = CIRCLED_DIGIT_RE.sub("", text)
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()\[\]{}·ㆍ,./‧、]", "", text)
    return text


def normalize_numeric_raw(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if text in MISSING_VALUE_TOKENS:
        return None
    text = text.replace(",", "").replace("%", "").strip()
    if text in MISSING_VALUE_TOKENS:
        return None
    return text


def parse_numeric(raw: Any) -> float | None:
    """문자열을 숫자로 변환합니다.

    성과지표 표(별첨1·별첨3·별첨6)에서는 값이 괄호로 싸여 있어도(예: "(4.11)")
    회계 문서의 음수 표기가 아니라 강조·잠정치 표시로 쓰인 사례가 실제
    확인됐습니다(달성률과 대조하면 양수가 맞음). 따라서 이 함수는 괄호를
    제거만 하고 부호는 원문의 `-`/`△` 표시가 있을 때만 음수로 판단합니다.
    """
    text = normalize_numeric_raw(raw)
    if text is None:
        return None
    text = text.strip()
    text = text.strip("()")
    negative = text.startswith(("-", "△"))
    text = text.lstrip("-△")
    try:
        value = float(text)
    except ValueError:
        return None
    return -value if negative else value


def clean_direction(raw: Any) -> tuple[str | None, bool]:
    """지표방향 원본값을 보존하고 `상향66:57`류 오염 여부를 별도로 표시합니다."""
    if raw is None:
        return None, False
    text = str(raw).strip()
    if text in ("상향", "하향", ""):
        return (text or None), False
    match = re.match(r"^(상향|하향)", text)
    if match and text != match.group(1):
        return text, True
    return text, False


def compute_achievement_rate(
    direction: str | None, target: float | None, actual: float | None
) -> float | None:
    """상향지표는 실적÷목표×100, 하향지표는 목표÷실적×100. 방향 불명/분모 0은 None."""
    if target is None or actual is None:
        return None
    normalized_direction = (direction or "").strip()
    if normalized_direction.startswith("상향"):
        if target == 0:
            return None
        return actual / target * 100
    if normalized_direction.startswith("하향"):
        if actual == 0:
            return None
        return target / actual * 100
    return None


def classify_numeric_match(
    manual_value: float | None,
    pdf_value: float | None,
    *,
    tolerance: float = 1e-6,
) -> str:
    """수기값과 PDF값을 비교합니다. 두 값 모두 있을 때만 정합 판정을 내립니다."""
    if manual_value is None and pdf_value is None:
        return "NOT_APPLICABLE"
    if manual_value is None and pdf_value is not None:
        return "PDF_MISSING_MANUAL_PRESENT" if False else "MANUAL_MISSING_PDF_PRESENT"
    if manual_value is not None and pdf_value is None:
        return "PDF_MISSING_MANUAL_PRESENT"
    diff = abs(manual_value - pdf_value)
    if diff <= tolerance:
        return "EXACT_MATCH"
    return "VALUE_MISMATCH"


def classify_rate_match(manual_rate: float | None, pdf_rate: float | None) -> str:
    """달성률은 반올림 오차(0.1%p 이내)와 실제 불일치를 분리합니다."""
    if manual_rate is None and pdf_rate is None:
        return "NOT_APPLICABLE"
    if manual_rate is None:
        return "MANUAL_MISSING_PDF_PRESENT"
    if pdf_rate is None:
        return "PDF_MISSING_MANUAL_PRESENT"
    diff = abs(manual_rate - pdf_rate)
    if diff <= 1e-9:
        return "EXACT_MATCH"
    if diff <= 0.1:
        return "ROUNDING_ONLY"
    return "VALUE_MISMATCH"


def numeric_change(
    baseline: float | None, comparison: float | None
) -> tuple[float | None, float | None]:
    """두 값의 절대 변화량과 baseline 대비 상대 변화율(%)만 계산합니다.

    왜 값이 바뀌었는지(정책 조정·오류·산식 차이 등)는 판단하지 않고, 순수
    산술 차이만 제공합니다. `VALUE_MISMATCH`/`RATE_MISMATCH`로 분류된 행이
    "얼마나" 다른지 사람이 63행을 일일이 손으로 빼지 않아도 되도록 하기
    위한 보조 컬럼입니다. baseline이 0이거나 둘 중 하나라도 없으면 상대
    변화율은 None입니다(절대 변화량은 baseline이 0이어도 계산됩니다).
    """
    if baseline is None or comparison is None:
        return None, None
    change_abs = comparison - baseline
    change_pct = (change_abs / abs(baseline) * 100) if baseline != 0 else None
    return change_abs, change_pct


# ---------------------------------------------------------------------------
# 위치→행 매핑이 있는 정규화 (원문 근거 역추적용)
# ---------------------------------------------------------------------------


def _normalize_with_line_map(lines: list[str]) -> tuple[str, list[int]]:
    norm_parts: list[str] = []
    line_map: list[int] = []
    for idx, line in enumerate(lines):
        norm_line = normalize_indicator_name(line)
        norm_parts.append(norm_line)
        line_map.extend([idx] * len(norm_line))
    return "".join(norm_parts), line_map


def _find_section(
    page_texts: list[str], start_marker: str, end_marker: str | None
) -> tuple[int, int] | None:
    """`start_marker`가 있는 페이지부터 `end_marker` 이전 페이지까지 0-based 구간."""
    start_idx = next((i for i, t in enumerate(page_texts) if start_marker in t), None)
    if start_idx is None:
        return None
    end_idx = len(page_texts) - 1
    if end_marker:
        for i in range(start_idx + 1, len(page_texts)):
            if end_marker in page_texts[i]:
                end_idx = i - 1
                break
    return start_idx, end_idx


# ---------------------------------------------------------------------------
# 보고서 별첨7: 성과지표 세부현황(목표치·실적치·달성률)
# ---------------------------------------------------------------------------


LABEL_TOKENS = ("목표", "실적", "달성률")
# 유니코드 개인용 영역(Private Use Area). 일부 정부 PDF는 임베드 폰트의 cmap이
# 불완전해 특정 한글 글자가 이 영역의 대체 코드로 추출됩니다. 이런 페이지는
# 텍스트 검색이 신뢰할 수 없으므로 OCR로 보강합니다.
PUA_CHAR_RE = re.compile(r"[\ue000-\uf8ff]")


@dataclass
class AchievementEvidence:
    matched_name: str
    split_pdf_page: int
    source_pdf_page: int
    printed_page: int | None
    source_text: str
    target_values_raw: list[str]
    actual_values_raw: list[str]
    rate_values_raw: list[str]
    extraction_method: str = "TEXT"
    source_file: str | None = None

    @property
    def target_raw(self) -> str | None:
        return self.target_values_raw[-1] if self.target_values_raw else None

    @property
    def actual_raw(self) -> str | None:
        return self.actual_values_raw[-1] if self.actual_values_raw else None

    @property
    def rate_raw(self) -> str | None:
        return self.rate_values_raw[-1] if self.rate_values_raw else None


def _collect_values_after_label(
    lines: list[str],
    start_line: int,
    label: str,
    *,
    max_scan: int = 40,
    max_values: int = 3,
) -> tuple[list[str], int]:
    """`label` 줄을 찾아 그 다음에 오는 연도별 값을 모읍니다.

    일부 페이지는 두 연도의 값이 한 줄에 공백으로 붙어 렌더링됩니다
    (예: "100.0% 144.3%"). 이런 줄을 하나의 값으로 처리하면 남은 값이 다음
    지표의 이름 줄까지 잘못 흡수합니다. 공백으로 나뉜 토큰이 모두 숫자/퍼센트
    형태면 여러 값으로 분리합니다.
    """
    n = len(lines)
    i = start_line
    while i < min(n, start_line + max_scan):
        compact = lines[i].strip().replace(" ", "")
        if compact == label or compact.startswith(f"{label}("):
            values: list[str] = []
            j = i + 1
            while j < n and len(values) < max_values:
                stripped = lines[j].strip()
                if stripped == "":
                    j += 1
                    continue
                if stripped in LABEL_TOKENS:
                    break
                tokens = stripped.split()
                if len(tokens) > 1 and all(STANDALONE_NUMERIC_RE.match(t) for t in tokens):
                    values.extend(tokens[: max_values - len(values)])
                else:
                    values.append(stripped)
                j += 1
            allowed_text = {
                "-",
                "(-)",
                "N/A",
                "신규",
                "집계중",
                "미산출",
                "해당없음",
                "미정",
                "X",
            }
            if values and all(
                parse_numeric(value) is not None or value.strip().upper() in allowed_text
                for value in values
            ):
                return values, j
        i += 1
    return [], start_line


def full_document_path(spec: PdfDocSpec) -> Path | None:
    """분리 별첨 파일명에 대응하는 원본 전체 PDF가 있으면 반환합니다."""
    full_name = re.sub(r"-\d+-\d+(?=\.pdf$)", "", spec.filename)
    path = PERFORMANCE_DOC_ROOT / full_name
    return path if path.is_file() else None


def _extract_text_achievement_evidence(
    pages: list[str],
    page_indexes: range,
    candidate_names: list[str],
    *,
    source_file: str,
    source_page: Callable[[int], int],
    require_complete_values: bool,
    extraction_method: str,
) -> dict[str, AchievementEvidence]:
    norm_lookup = {normalize_indicator_name(name): name for name in candidate_names}
    results: dict[str, AchievementEvidence] = {}
    for page_offset in page_indexes:
        raw_text = pages[page_offset]
        lines = raw_text.split("\n")
        flat_norm, line_map = _normalize_with_line_map(lines)
        matches: list[tuple[int, int, str]] = []
        for norm_name, original_name in norm_lookup.items():
            if not norm_name or original_name in results:
                continue
            cursor = 0
            while (start := flat_norm.find(norm_name, cursor)) != -1:
                matches.append((start, start + len(norm_name), original_name))
                cursor = start + len(norm_name)
        matches.sort(key=lambda item: item[0])

        for start, end, original_name in matches:
            line_end = line_map[end - 1] if end > 0 else 0
            year_count = len(set(re.findall(r"['’]?(\d{2})년", raw_text)))
            max_values = min(4, max(3, year_count))
            target_values, cursor = _collect_values_after_label(
                lines, line_end, "목표", max_values=max_values
            )
            actual_values, cursor = _collect_values_after_label(
                lines, cursor, "실적", max_values=max_values
            )
            rate_values, cursor = _collect_values_after_label(
                lines, cursor, "달성률", max_values=max_values
            )
            if require_complete_values and not (target_values and actual_values and rate_values):
                continue
            source_text = "\n".join(
                line for line in lines[max(line_end - 1, 0) : cursor] if line.strip()
            )[:600]
            results[original_name] = AchievementEvidence(
                matched_name=original_name,
                split_pdf_page=page_offset + 1,
                source_pdf_page=source_page(page_offset + 1),
                printed_page=printed_page_number(raw_text),
                source_text=source_text,
                target_values_raw=target_values,
                actual_values_raw=actual_values,
                rate_values_raw=rate_values,
                extraction_method=extraction_method,
                source_file=source_file,
            )
    return results


def extract_report_achievement_evidence(
    report_spec: PdfDocSpec, indicator_names: list[str]
) -> dict[str, AchievementEvidence]:
    """별첨3 "3. 세부현황"에서 지표명별 목표·실적·달성률(최근 3개년)을 찾습니다.

    이 표는 "목표"/"실적"/"달성률" 세 줄 라벨 아래 최근 3개년 값을 순서대로
    나열하므로, 라벨을 앵커로 사용해 위치 기반 추정보다 신뢰도 높게
    추출합니다. 세 값 중 마지막 값이 해당 보고서의 회계연도 값입니다.
    """
    pages = load_page_texts(report_spec.path)
    section = _find_section(pages, "별첨3", "별첨4")
    if section is None:
        raise PdfReconciliationError(f"별첨3 구간을 찾을 수 없습니다: {report_spec.path}")
    start_idx, end_idx = section

    candidate_names = sorted(
        {name for name in indicator_names if name},
        key=lambda n: -len(normalize_indicator_name(n)),
    )
    source_filename = getattr(report_spec, "filename", report_spec.path.name)
    results = _extract_text_achievement_evidence(
        pages,
        range(start_idx, end_idx + 1),
        candidate_names,
        source_file=source_filename,
        source_page=report_spec.source_pdf_page,
        require_complete_values=False,
        extraction_method="TEXT",
    )

    # 별첨3은 부처에 따라 지표명만 싣고 목표·실적·달성률 상세표는 본문 앞부분에
    # 둡니다. 같은 파일의 전체본이 있으면 별첨 시작 전 본문에서 세 값이 모두
    # 확인되는 근거만 채택해, 지표명만 잡힌 빈 별첨 근거를 대체합니다.
    full_path = full_document_path(report_spec) if isinstance(report_spec, PdfDocSpec) else None
    if full_path is not None:
        full_pages = load_page_texts(full_path)
        detail_end = min(len(full_pages), report_spec.source_page_start - 1)
        results.update(
            _extract_text_achievement_evidence(
                full_pages,
                range(detail_end),
                candidate_names,
                source_file=full_path.name,
                source_page=lambda page: page,
                require_complete_values=True,
                extraction_method="FULL_TEXT",
            )
        )

    # 텍스트 검색으로 찾지 못한 지표는, 해당 구간에 PUA(개인용 영역) 글자가
    # 있는 페이지에서만 OCR로 재시도합니다. 표 구조가 OCR에서 깨지므로
    # 목표·실적·달성률 3분류 라벨 앵커 없이 원문 창만 확보하고, 호출부에서
    # `ocr_status`를 `OCR_REQUIRED`로 표시해 사람 검토를 요구해야 합니다.
    still_missing = {name for name in candidate_names if name not in results}
    if still_missing:
        for page_offset in range(start_idx, end_idx + 1):
            if not still_missing:
                break
            raw_text = pages[page_offset]
            if not PUA_CHAR_RE.search(raw_text):
                continue
            try:
                ocr_text = ocr_page_text(report_spec.path, page_offset)
            except Exception:  # noqa: BLE001, S112 - Tesseract 미설치·페이지 렌더 실패는
                # 실패한 페이지 하나만 건너뛰고 나머지 지표 매칭을 계속합니다.
                # 이 지표들은 결국 매칭되지 않으면 OCR_REQUIRED로 표시되므로
                # 조용히 삭제되지 않습니다.
                continue
            ocr_norm = normalize_indicator_name(ocr_text)
            for original_name in list(still_missing):
                norm_name = normalize_indicator_name(original_name)
                # OCR 오인식 여지가 있으므로 이름 앞부분(6자 이상)만 대략 대조합니다.
                probe = norm_name[: max(6, len(norm_name) // 2)]
                if probe and probe in ocr_norm:
                    results[original_name] = AchievementEvidence(
                        matched_name=original_name,
                        split_pdf_page=page_offset + 1,
                        source_pdf_page=report_spec.source_pdf_page(page_offset + 1),
                        printed_page=printed_page_number(raw_text),
                        source_text=ocr_text.strip()[:600],
                        target_values_raw=[],
                        actual_values_raw=[],
                        rate_values_raw=[],
                        extraction_method="OCR",
                        source_file=source_filename,
                    )
                    still_missing.discard(original_name)
    return results


# ---------------------------------------------------------------------------
# 보고서 별첨6: 성과계획서 변경 사항 (계획서 확정값 교차검증)
# ---------------------------------------------------------------------------
#
# 표 구조: 각 지표 행은 "예산전\n예산후\n지표명(전)\n가중치(전)\n지표명(후)\n
# 가중치(후)\n목표치(전)\n목표치(후)\n분야\n<수정사유>" 순서입니다. 지표명이
# 여러 줄로 줄바꿈되고 원문자 숫자·병합 셀이 섞여 있어 행 전체를 하나의
# 정규식으로 안정적으로 파싱하기 어렵습니다. 대신 지표명을 앵커로 검색해
# 그 지점부터 다음 "<사유>" 토큰까지를 그 지표의 변경 근거 창으로 사용합니다.
# 이 창의 마지막 숫자가 목표치(후, 확정값)입니다. 지표명이 실제로 바뀐 경우
# name_before와 name_after가 창 안에서 다르게 나타납니다.

REASON_TOKEN_RE = re.compile(r"<([^<>]{1,40})>")
# "수정사유" 열이 <괄호> 없이 자유 서술문(예: "'22년 실적치 고려하여...")으로
# 적힌 경우, 다음 지표 행의 "예산 변경" 숫자(쉼표 포함 순수 숫자 줄)가 나오면
# 그 지표의 서술이 끝난 것으로 보고 멈춥니다.
BUDGET_LINE_RE = re.compile(r"^[\d,]{4,}$")
# 표의 "분야" 열에 나오는 고정 범주어. 이 줄이 나오면 그 앞 숫자가 목표치(후)이고,
# 그 뒤는 수정 사유 서술(때로는 <>로 감싸지 않은 긴 문장)이라 숫자 후보에서
# 제외해야 합니다.
CATEGORY_LINE_TOKENS = frozenset(
    {
        "일반",
        "재정",
        "R&D",
        "기금",
        "보조",
        "출자",
        "융자",
        "보증",
        "기타",
        "정보",
        "화",
        "정보화",
        "특별회계",
        "일반회계",
        "공제",
    }
)


@dataclass
class ChangeEvidence:
    matched_name: str
    window_text: str
    target_before_raw: str | None
    target_after_raw: str | None
    reason: str | None
    split_pdf_page: int
    source_pdf_page: int
    printed_page: int | None
    source_text: str


def _extract_free_text_reason(lines: list[str], start_line: int, max_lines: int = 8) -> str | None:
    """분야(category) 줄 다음의 자유 서술 수정사유를 모읍니다.

    `<괄호>` 사유가 없을 때만 호출됩니다. 다음 지표 행의 예산 숫자나
    "프로그램목표"/"전략목표" 헤더가 나오면 이 지표의 서술이 끝난 것으로
    보고 멈춥니다. 확정할 수 없으면(경계를 못 찾으면) `max_lines`에서
    안전하게 끊습니다.
    """
    collected: list[str] = []
    for k in range(start_line, min(len(lines), start_line + max_lines)):
        text = lines[k].strip()
        if not text:
            continue
        if text in CATEGORY_LINE_TOKENS:
            continue
        if BUDGET_LINE_RE.match(text.replace(" ", "")):
            break
        if text.startswith(("프로그램목표", "전략목표")):
            break
        collected.append(text)
    return " ".join(collected).strip() or None


def find_change_evidence(report_spec: PdfDocSpec, indicator_name: str) -> ChangeEvidence | None:
    """별첨6에서 `indicator_name`(변경 전 또는 후 명칭 모두 허용)을 찾습니다."""
    norm_name = normalize_indicator_name(indicator_name)
    if not norm_name:
        return None
    pages = load_page_texts(report_spec.path)
    section = _find_section(pages, "별첨6", "별첨7")
    if section is None:
        return None
    start_idx, end_idx = section

    for page_offset in range(start_idx, end_idx + 1):
        raw_text = pages[page_offset]
        lines = raw_text.split("\n")
        flat_norm, line_map = _normalize_with_line_map(lines)
        pos = flat_norm.find(norm_name)
        if pos == -1:
            continue
        end_pos = pos + len(norm_name)
        line_start = line_map[pos]
        line_after_match = line_map[end_pos - 1] if end_pos > 0 else line_start
        # 다음 "<사유>" 토큰과 "분야"(일반/재정/R&D 등 고정 범주어) 줄 중 먼저
        # 나오는 지점까지를 이 지표 행의 나머지로 봅니다. 사유가 <>로 감싸지
        # 않은 긴 서술문일 때 범주어 경계가 없으면 숫자 후보가 사유 문장의
        # 다른 숫자(예산액 등)까지 잘못 흡수합니다.
        reason_line = None
        category_line = None
        for j in range(line_after_match, min(len(lines), line_after_match + 25)):
            stripped_j = lines[j].strip()
            if (
                category_line is not None
                and reason_line is None
                and BUDGET_LINE_RE.match(stripped_j.replace(" ", ""))
            ):
                # 이 지표의 "분야" 줄까지는 찾았는데 <사유> 토큰이 아직 없는
                # 상태에서 다음 지표 행의 예산(변경전/변경후) 숫자 줄이 나오면,
                # 이 지표의 수정사유 칸은 원문에서 실제로 빈 칸입니다. 계속
                # 찾으면 다음 지표의 <사유>를 이 지표의 사유로 잘못 흡수하게
                # 되므로 여기서 멈추고 reason_line=None을 유지합니다.
                break
            if reason_line is None and REASON_TOKEN_RE.search(lines[j]):
                reason_line = j
            if category_line is None and stripped_j in CATEGORY_LINE_TOKENS:
                category_line = j
            if reason_line is not None and category_line is not None:
                break
        context_end_line = (
            reason_line if reason_line is not None else (category_line or line_after_match) + 4
        )

        reason_match = REASON_TOKEN_RE.search(
            "\n".join(lines[max(line_start - 1, 0) : reason_line + 1])
            if reason_line is not None
            else ""
        )
        reason = reason_match.group(1).strip() if reason_match else None

        # "분야" 범주어 줄(들)이 끝난 다음부터가 자유 서술 사유 후보입니다.
        # `<괄호>` 사유를 못 찾았을 때만 자유 서술을 시도합니다.
        free_reason_end_line = context_end_line
        if reason is None and category_line is not None:
            free_start = category_line + 1
            while free_start < len(lines) and lines[free_start].strip() in CATEGORY_LINE_TOKENS:
                free_start += 1
            reason = _extract_free_text_reason(lines, free_start)
            if reason:
                free_reason_end_line = max(context_end_line, free_start + 4)

        window_lines = lines[max(line_start - 1, 0) : free_reason_end_line + 1]
        window_text = "\n".join(ln for ln in window_lines if ln.strip())

        numbers_end_line = category_line if category_line is not None else context_end_line
        numbers_text = "\n".join(lines[max(line_start - 1, 0) : numbers_end_line])
        numbers = [
            tok for tok in NUMBER_TOKEN_RE.findall(numbers_text) if parse_numeric(tok) is not None
        ]
        target_after_raw = numbers[-1] if numbers else None
        target_before_raw = numbers[-2] if len(numbers) >= 2 else None

        return ChangeEvidence(
            matched_name=indicator_name,
            window_text=window_text[:500],
            target_before_raw=target_before_raw,
            target_after_raw=target_after_raw,
            reason=reason,
            split_pdf_page=page_offset + 1,
            source_pdf_page=report_spec.source_pdf_page(page_offset + 1),
            printed_page=printed_page_number(raw_text),
            source_text=window_text[:500],
        )
    return None


# ---------------------------------------------------------------------------
# 계획서 별첨1: 프로그램 성과지표 현황 (원문 근거·페이지 인용)
# ---------------------------------------------------------------------------


@dataclass
class PlanEvidence:
    matched_name: str
    split_pdf_page: int
    source_pdf_page: int
    printed_page: int | None
    source_text: str
    extraction_method: str
    numeric_candidates: list[float]


def extract_plan_evidence(
    plan_spec: PdfDocSpec,
    indicator_names: list[str],
    *,
    image_only: bool | None = None,
    max_page_count: int | None = None,
) -> dict[str, PlanEvidence]:
    """별첨1(세부현황)에서 지표명별 원문 근거와 목표치 후보를 찾습니다.

    2023·2024년은 실제 텍스트(블록 단위)를 사용합니다. 2022년은 표 셀이
    이미지이므로 페이지를 렌더링해 로컬 OCR을 수행합니다. OCR 결과는 참고용
    후보이며, 최종 상태는 이 함수를 호출하는 쪽에서 `ocr_status`를
    `OCR_REQUIRED`로 별도 표시해 사람 검토를 유도합니다.
    """
    pages = load_page_texts(plan_spec.path)
    # 분리 PDF 파일 자체가 별첨1부터 시작하도록 잘려 있습니다. 연도별로 표지에
    # "별첨1" 글자가 이미지로만 있거나(2023) 텍스트로 있거나(2024) 달라
    # 마커 탐색이 불안정하므로, 항상 1페이지를 시작으로 두고 "별첨2" 텍스트가
    # 나오는 페이지 전까지만 사용합니다.
    start_idx = 0
    end_idx = len(pages) - 1
    if max_page_count is not None:
        end_idx = min(end_idx, max_page_count - 1)
    for i in range(1, len(pages)):
        if "별첨2" in pages[i]:
            end_idx = i - 1
            break

    candidate_names = sorted(
        {name for name in indicator_names if name},
        key=lambda n: -len(normalize_indicator_name(n)),
    )
    norm_lookup = {normalize_indicator_name(n): n for n in candidate_names}
    results: dict[str, PlanEvidence] = {}
    is_image_only = (
        PLAN_TABLE_IS_IMAGE_ONLY.get(plan_spec.fiscal_year, True)
        if image_only is None
        else image_only
    )

    with fitz.open(plan_spec.path) as document:
        for page_offset in range(start_idx, end_idx + 1):
            raw_text = pages[page_offset]
            printed = printed_page_number(raw_text)

            if is_image_only:
                try:
                    ocr_text = ocr_page_text(plan_spec.path, page_offset)
                except Exception:  # noqa: BLE001
                    ocr_text = ""
                # 텍스트 레이어가 깨진 PDF는 위의 별첨2 탐색이 실패합니다.
                # OCR에서 별첨2가 보이면 그 페이지부터는 별첨1 근거가 아니므로 중단합니다.
                if page_offset > 0 and "별첨2" in normalize_indicator_name(ocr_text):
                    break
                blocks_text = [ocr_text]
                extraction_method = "OCR"
            else:
                blocks_text = [b[4] for b in document[page_offset].get_text("blocks")]
                extraction_method = "TEXT"

            # 표 셀이 블록 단위로 잘게 나뉘어(예: 지표명이 목표치·단위 값과
            # 다른 블록으로 분리) 지표명이 한 블록 안에 온전히 들어있지 않을 수
            # 있습니다. 페이지의 모든 블록을 읽기 순서로 이어붙여 지표명을
            # 블록 경계를 넘어 찾고, 매칭된 지점부터 이후 몇 개 블록까지를
            # 근거 창으로 사용합니다.
            norm_blocks = [normalize_indicator_name(b) for b in blocks_text]
            block_starts: list[int] = []
            cursor = 0
            for nb in norm_blocks:
                block_starts.append(cursor)
                cursor += len(nb)
            flat_norm = "".join(norm_blocks)

            for norm_name, original_name in list(norm_lookup.items()):
                if not norm_name or original_name in results:
                    continue
                pos = flat_norm.find(norm_name)
                if pos == -1:
                    continue
                end_pos = pos + len(norm_name)
                start_block = (
                    next(
                        i for i, s in enumerate(block_starts) if s <= pos < s + len(norm_blocks[i])
                    )
                    if any(s <= pos < s + len(norm_blocks[i]) for i, s in enumerate(block_starts))
                    else 0
                )
                end_block = start_block
                for i, s in enumerate(block_starts):
                    if s < end_pos:
                        end_block = i
                window_blocks = blocks_text[start_block : end_block + 4]
                window_text = "".join(window_blocks)
                numbers = [
                    v
                    for v in (parse_numeric(tok) for tok in NUMBER_TOKEN_RE.findall(window_text))
                    if v is not None
                ]
                results[original_name] = PlanEvidence(
                    matched_name=original_name,
                    split_pdf_page=page_offset + 1,
                    source_pdf_page=plan_spec.source_pdf_page(page_offset + 1),
                    printed_page=printed,
                    source_text=re.sub(r"\s+", " ", window_text).strip()[:400],
                    extraction_method=extraction_method,
                    numeric_candidates=numbers,
                )
    return results


def select_plan_target(evidence: PlanEvidence, manual_target: float | None) -> float | None:
    if not evidence.numeric_candidates:
        return None
    if manual_target is None:
        # 마지막 숫자가 목표치일 가능성이 높다는 열 순서 기반 보수적 추정.
        return evidence.numeric_candidates[-1]
    return min(evidence.numeric_candidates, key=lambda x: abs(x - manual_target))


# ---------------------------------------------------------------------------
# 63행 전체 대조 오케스트레이션
# ---------------------------------------------------------------------------

REQUIRED_MANUAL_COLUMNS: tuple[str, ...] = (
    "source_indicator_id",
    "ministry_name",
    "fiscal_year",
    "strategic_goal_number",
    "program_goal_number",
    "source_program_code",
    "performance_program_name",
    "indicator_name_plan",
    "indicator_name_report",
    "indicator_unit",
    "indicator_direction",
    "planned_target_raw",
    "actual_value_raw",
    "official_achievement_rate_raw",
    "planned_target_numeric",
    "actual_value_numeric",
    "official_achievement_rate_numeric",
    "plan_report_match_status_raw",
)


def _pick_overall_status(statuses: list[str]) -> str:
    present = [s for s in statuses if s in ALLOWED_STATUS_VALUES]
    non_na = [s for s in present if s != "NOT_APPLICABLE"]
    pool = non_na or present or ["NOT_APPLICABLE"]
    for candidate in STATUS_PRIORITY:
        if candidate in pool:
            return candidate
    return "MANUAL_REVIEW"


def _name_equal(a: str | None, b: str | None) -> bool:
    return bool(a) and bool(b) and normalize_indicator_name(a) == normalize_indicator_name(b)


def report_section_has_pua_glyphs(report_spec: PdfDocSpec) -> bool:
    """별첨3 구간에 폰트 매핑이 불완전한 개인용 영역(PUA) 글자가 있는지 확인합니다.

    있으면 텍스트 검색으로 못 찾은 지표도 `PDF_NOT_FOUND`로 단정하지 않고
    `OCR_REQUIRED`로 사람 확인을 유도해야 합니다.
    """
    pages = load_page_texts(report_spec.path)
    section = _find_section(pages, "별첨3", "별첨4")
    if section is None:
        return False
    start_idx, end_idx = section
    return any(PUA_CHAR_RE.search(pages[i]) for i in range(start_idx, end_idx + 1))


def reconcile_row(
    row: dict[str, Any],
    *,
    plan_evidence_by_year: dict[int, dict[str, PlanEvidence]],
    report_evidence_by_year: dict[int, dict[str, AchievementEvidence]],
    plan_specs_valid_pages: dict[int, tuple[int, int]],
    report_specs_valid_pages: dict[int, tuple[int, int]],
    report_has_pua_by_year: dict[int, bool] | None = None,
    report_structure_review_by_year: dict[int, bool] | None = None,
    plan_image_only_by_year: dict[int, bool] | None = None,
    doc_specs: tuple[PdfDocSpec, ...] = PDF_DOC_SPECS,
) -> dict[str, Any]:
    """수기 63행 중 한 행을 PDF 근거와 대조해 최종 스키마 한 행을 만듭니다."""
    year = int(row["fiscal_year"])
    plan_spec = doc_spec(year, "plan", doc_specs)
    report_spec = doc_spec(year, "report", doc_specs)

    manual_target = row.get("planned_target_numeric")
    manual_actual = row.get("actual_value_numeric")
    manual_rate = row.get("official_achievement_rate_numeric")
    manual_target = None if pd.isna(manual_target) else float(manual_target)
    manual_actual = None if pd.isna(manual_actual) else float(manual_actual)
    manual_rate = None if pd.isna(manual_rate) else float(manual_rate)

    direction_raw = row.get("indicator_direction")
    direction_clean, direction_polluted = clean_direction(direction_raw)

    review_reasons: list[str] = []
    if direction_polluted:
        review_reasons.append("DIRECTION_PARSE_REVIEW")

    # --- 계획서 별첨1 (필요 시 별첨6 확정값으로 보강) --------------------
    plan_name = row.get("indicator_name_plan")
    plan_ev_map = plan_evidence_by_year.get(year, {})
    plan_ev = plan_ev_map.get(plan_name)

    plan_change_ev = find_change_evidence(report_spec, plan_name) if plan_name else None
    if plan_change_ev is None and row.get("indicator_name_report"):
        plan_change_ev = find_change_evidence(report_spec, row["indicator_name_report"])

    is_plan_image_only = (plan_image_only_by_year or PLAN_TABLE_IS_IMAGE_ONLY).get(year, True)
    plan_target_source = "NONE"
    pdf_plan_target_raw: str | None = None
    pdf_plan_indicator_name: str | None = None
    plan_source_text: str | None = None
    plan_split_page: int | None = None
    plan_source_page: int | None = None
    plan_printed_page: int | None = None
    plan_extraction_method = "NONE"

    if plan_ev is not None:
        pdf_plan_indicator_name = plan_ev.matched_name
        pdf_plan_target_raw_val = select_plan_target(plan_ev, manual_target)
        pdf_plan_target_raw = (
            None if pdf_plan_target_raw_val is None else str(pdf_plan_target_raw_val)
        )
        plan_source_text = plan_ev.source_text
        plan_split_page = plan_ev.split_pdf_page
        plan_source_page = plan_ev.source_pdf_page
        plan_printed_page = plan_ev.printed_page
        plan_extraction_method = plan_ev.extraction_method
        plan_target_source = "별첨1"

    if pdf_plan_target_raw is None and plan_change_ev is not None:
        pdf_plan_indicator_name = pdf_plan_indicator_name or plan_change_ev.matched_name
        pdf_plan_target_raw = plan_change_ev.target_after_raw
        plan_extraction_method = plan_extraction_method if plan_ev is not None else "TEXT"
        plan_target_source = "별첨1+별첨6" if plan_ev is not None else "별첨6만"
        review_reasons.append(
            "PLAN_TARGET_FROM_CHANGE_TABLE_ONLY"
            if plan_ev is None
            else "PLAN_TARGET_CROSSCHECKED_WITH_CHANGE_TABLE"
        )

    planned_target_numeric_pdf = parse_numeric(pdf_plan_target_raw)

    # --- 별첨6 "성과계획서 변경 사항" 표: 계획-보고 목표가 왜 바뀌었는지에
    # 대한 공식 문서 근거. `plan_ev`(별첨1) 값이 이미 있어도(OCR 포함) 항상
    # 별도 컬럼으로 노출합니다. 원본 추출값(`pdf_plan_target_raw`)은 그대로
    # 보존하고 덮어쓰지 않으며, 이 표가 OCR값과 다르면 사람이 봐야 할
    # 근거로만 플래그를 남깁니다(자동으로 값을 바꾸지 않음).
    documented_change_target_before_raw = (
        plan_change_ev.target_before_raw if plan_change_ev is not None else None
    )
    documented_change_target_after_raw = (
        plan_change_ev.target_after_raw if plan_change_ev is not None else None
    )
    documented_change_reason_raw = plan_change_ev.reason if plan_change_ev is not None else None
    # 별첨6 페이지는 인쇄 쪽번호("- NN -") 자체가 없는 경우가 많아
    # (실제 데이터 확인 결과 63행 전부 `printed_page`가 None), 항상 채워지는
    # 분리 PDF·원본 PDF 페이지 번호를 근거 추적용으로 씁니다. 인쇄 쪽번호가
    # 있으면 참고용으로 별도 보존합니다.
    documented_change_split_pdf_page = (
        plan_change_ev.split_pdf_page if plan_change_ev is not None else None
    )
    documented_change_source_file = report_spec.filename if plan_change_ev is not None else None
    documented_change_source_pdf_page = (
        plan_change_ev.source_pdf_page if plan_change_ev is not None else None
    )
    documented_change_printed_page = (
        plan_change_ev.printed_page if plan_change_ev is not None else None
    )
    documented_change_target_before_numeric = parse_numeric(documented_change_target_before_raw)
    documented_change_target_after_numeric = parse_numeric(documented_change_target_after_raw)
    documented_change_ocr_conflict = (
        plan_extraction_method == "OCR"
        and documented_change_target_before_numeric is not None
        and planned_target_numeric_pdf is not None
        and abs(documented_change_target_before_numeric - planned_target_numeric_pdf) > 1e-6
    )
    if documented_change_ocr_conflict:
        review_reasons.append("OCR_TARGET_CONTRADICTS_CHANGE_TABLE")
    if documented_change_reason_raw is not None:
        review_reasons.append("TARGET_CHANGE_REASON_DOCUMENTED_IN_별첨6")

    if plan_ev is None and plan_change_ev is None:
        plan_name_match_status = "PDF_NOT_FOUND"
        plan_target_match_status = "PDF_NOT_FOUND"
    else:
        plan_name_match_status = "EXACT_MATCH" if plan_ev is not None else "MATCH_AFTER_CHANGE"
        plan_target_match_status = classify_numeric_match(manual_target, planned_target_numeric_pdf)

    # 2022년 별첨1은 이미지 표라 원문 자체 확인에는 항상 OCR이 필요합니다.
    # 별첨6 텍스트로 확정값을 교차검증했더라도, 계획서 원문 자체의 자동
    # 확인은 아니므로 OCR_REQUIRED를 유지합니다.
    if is_plan_image_only or plan_extraction_method == "OCR":
        plan_ocr_status = "OCR_REQUIRED"
    else:
        plan_ocr_status = "NOT_APPLICABLE"

    # --- 보고서 별첨3 "3.세부현황" ---------------------------------------
    report_name = row.get("indicator_name_report")
    report_ev_map = report_evidence_by_year.get(year, {})
    report_ev = report_ev_map.get(report_name)

    pdf_report_target_raw: str | None = None
    pdf_report_actual_raw: str | None = None
    pdf_report_rate_raw: str | None = None
    pdf_report_indicator_name: str | None = None
    report_source_text: str | None = None
    report_source_file = report_spec.filename
    report_split_page: int | None = None
    report_source_page: int | None = None
    report_printed_page: int | None = None
    report_extraction_method = "NONE"

    if report_ev is not None:
        pdf_report_indicator_name = report_ev.matched_name
        pdf_report_target_raw = report_ev.target_raw
        pdf_report_actual_raw = report_ev.actual_raw
        pdf_report_rate_raw = report_ev.rate_raw
        report_source_text = report_ev.source_text
        report_split_page = report_ev.split_pdf_page
        report_source_page = report_ev.source_pdf_page
        report_printed_page = report_ev.printed_page
        report_extraction_method = report_ev.extraction_method
        report_source_file = report_ev.source_file or report_spec.filename

    actual_value_numeric_pdf = parse_numeric(pdf_report_actual_raw)
    report_target_numeric_pdf = parse_numeric(pdf_report_target_raw)
    official_rate_numeric_pdf = parse_numeric(pdf_report_rate_raw)

    # 별첨6 "변경 후" 목표(문서상 최종 확정값)와 별첨3에서 직접 추출한 보고서
    # 목표가 서로 다르면 두 근거 문서가 어긋난다는 뜻이라 사람이 봐야 합니다.
    # (자동으로 어느 쪽이 맞는지 결정하지 않습니다.)
    documented_change_report_conflict = (
        documented_change_target_after_numeric is not None
        and report_target_numeric_pdf is not None
        and abs(documented_change_target_after_numeric - report_target_numeric_pdf) > 1e-6
    )
    if documented_change_report_conflict:
        review_reasons.append("CHANGE_TABLE_TARGET_CONTRADICTS_REPORT_TABLE")

    if report_ev is None:
        structure_review = (report_structure_review_by_year or {}).get(year, False)
        # 텍스트 검색으로 못 찾았어도, 이 연도 별첨3 구간에 폰트 매핑이 불완전한
        # PUA 글자가 있으면 "지표가 없다"가 아니라 "OCR로 사람이 확인해야
        # 한다"가 맞는 판정입니다.
        has_pua = (report_has_pua_by_year or {}).get(year, False)
        if structure_review:
            report_name_match_status = "MANUAL_REVIEW"
            report_target_match_status = "MANUAL_REVIEW"
            report_actual_match_status = "MANUAL_REVIEW"
            report_achievement_rate_match_status = "MANUAL_REVIEW"
            report_ocr_status = "NOT_APPLICABLE"
            review_reasons.append("REPORT_APPENDIX_STRUCTURE_REVIEW")
        elif has_pua:
            report_name_match_status = "OCR_REQUIRED"
            report_target_match_status = "OCR_REQUIRED"
            report_actual_match_status = "OCR_REQUIRED"
            report_achievement_rate_match_status = "OCR_REQUIRED"
            report_ocr_status = "OCR_REQUIRED"
            review_reasons.append("REPORT_TEXT_LAYER_PUA_GLYPH_OCR_NEEDED")
        else:
            report_name_match_status = "PDF_NOT_FOUND"
            report_target_match_status = "PDF_NOT_FOUND"
            report_actual_match_status = "PDF_NOT_FOUND"
            report_achievement_rate_match_status = "PDF_NOT_FOUND"
            report_ocr_status = "NOT_APPLICABLE"
    else:
        report_name_match_status = "EXACT_MATCH"
        report_target_match_status = classify_numeric_match(
            manual_target, report_target_numeric_pdf
        )
        # 계획서 원문 목표치(수기값)와 보고서 자체 표의 목표치가 다르면 1차로는
        # VALUE_MISMATCH지만, 별첨6 "성과계획서 변경 사항"표가 정확히 같은
        # 변경전→변경후 값을 문서화하고 있으면 이는 추출 오류가 아니라 목표치
        # 사후 개정입니다(실적 확정 반영, 기재부 검토 반영, 단순 오기 정정 등
        # 사유는 documented_change_reason_raw에 별도 보존). 이 경우에만
        # MATCH_AFTER_CHANGE로 재분류합니다 — 명칭 유사도가 아니라 공식 변경표
        # 숫자가 정확히 일치할 때만 적용되므로 임의 확정이 아닙니다.
        if (
            report_target_match_status == "VALUE_MISMATCH"
            and documented_change_target_before_numeric is not None
            and documented_change_target_after_numeric is not None
            and manual_target is not None
            and report_target_numeric_pdf is not None
            and abs(documented_change_target_before_numeric - manual_target) <= 1e-6
            and abs(documented_change_target_after_numeric - report_target_numeric_pdf) <= 1e-6
        ):
            report_target_match_status = "MATCH_AFTER_CHANGE"
            review_reasons.append("REPORT_TARGET_CHANGE_CONFIRMED_BY_별첨6")
        report_actual_match_status = classify_numeric_match(manual_actual, actual_value_numeric_pdf)
        report_achievement_rate_match_status = classify_rate_match(
            manual_rate, official_rate_numeric_pdf
        )
        report_ocr_status = (
            "OCR_REQUIRED" if report_extraction_method == "OCR" else "NOT_APPLICABLE"
        )
        if report_extraction_method == "OCR":
            # OCR 경로는 표 라벨(목표/실적/달성률) 앵커를 쓰지 못해 숫자 후보가
            # 없습니다. 사람이 원문 이미지를 봐야 확정할 수 있습니다.
            report_target_match_status = "OCR_REQUIRED"
            report_actual_match_status = "OCR_REQUIRED"
            report_achievement_rate_match_status = "OCR_REQUIRED"

    ocr_status = _pick_overall_status([plan_ocr_status, report_ocr_status])
    if ocr_status == "NOT_APPLICABLE" and (
        plan_ocr_status == "OCR_REQUIRED" or report_ocr_status == "OCR_REQUIRED"
    ):
        ocr_status = "OCR_REQUIRED"

    computed_achievement_rate = compute_achievement_rate(
        direction_clean, report_target_numeric_pdf, actual_value_numeric_pdf
    )

    # --- 값 차이(변경량) 계산: "왜 바뀌었는지"가 아니라 "얼마나 바뀌었는지"만
    # 순수 산술로 계산합니다. 사람 검토가 필요한 판단(변경 사유)과 분리합니다.
    plan_target_change_abs, plan_target_change_pct = numeric_change(
        manual_target, planned_target_numeric_pdf
    )
    report_target_change_abs, report_target_change_pct = numeric_change(
        manual_target, report_target_numeric_pdf
    )
    report_actual_change_abs, report_actual_change_pct = numeric_change(
        manual_actual, actual_value_numeric_pdf
    )
    report_achievement_rate_change_abs_pp, _ = numeric_change(
        manual_rate, official_rate_numeric_pdf
    )
    # 계획서 목표치 원문과 보고서 목표치 원문 자체의 차이(둘 다 PDF값).
    # 이 값이 수기 데이터의 결측·오염 여부와 무관하게 "계획-보고 목표가
    # 얼마나 바뀌었는지"에 대한 직접적인 답입니다.
    plan_vs_report_target_change_abs, plan_vs_report_target_change_pct = numeric_change(
        planned_target_numeric_pdf, report_target_numeric_pdf
    )

    # --- 계획서-보고서 지표명 변경 여부 -----------------------------------
    names_differ = not _name_equal(plan_name, report_name)
    if names_differ:
        if plan_change_ev is not None:
            review_reasons.append("PLAN_REPORT_NAME_CHANGE_DOCUMENTED_IN_별첨6")
        else:
            review_reasons.append("PLAN_REPORT_NAME_CHANGE_UNDOCUMENTED")

    # --- 페이지 근거 상태 --------------------------------------------------
    plan_lo, plan_hi = plan_specs_valid_pages.get(year, (1, 10**6))
    report_lo, report_hi = report_specs_valid_pages.get(year, (1, 10**6))
    plan_evidence_source_page = plan_source_page or documented_change_source_pdf_page
    plan_page_ok = (
        plan_lo <= plan_source_page <= plan_hi
        if plan_source_page is not None
        else documented_change_source_pdf_page is None
        or report_lo <= documented_change_source_pdf_page <= report_hi
    )
    report_page_ok = report_source_page is None or report_lo <= report_source_page <= report_hi
    if plan_evidence_source_page is None and report_source_page is None:
        page_evidence_status = "PDF_NOT_FOUND"
    elif not plan_page_ok or not report_page_ok:
        page_evidence_status = "MANUAL_REVIEW"
        review_reasons.append("PAGE_OUT_OF_RANGE")
    elif plan_evidence_source_page is None or report_source_page is None:
        page_evidence_status = (
            "MANUAL_MISSING_PDF_PRESENT" if False else "PDF_MISSING_MANUAL_PRESENT"
        )
    else:
        page_evidence_status = "EXACT_MATCH"

    overall_status = _pick_overall_status(
        [
            plan_name_match_status,
            plan_target_match_status,
            report_name_match_status,
            report_target_match_status,
            report_actual_match_status,
            report_achievement_rate_match_status,
            ocr_status,
        ]
    )
    if names_differ and not plan_change_ev and overall_status in ("EXACT_MATCH", "ROUNDING_ONLY"):
        overall_status = "MANUAL_REVIEW"
        review_reasons.append("NAME_CHANGE_OVERRIDES_NUMERIC_MATCH")

    return {
        "source_indicator_id": row.get("source_indicator_id"),
        "ministry_code": plan_spec.ministry_code,
        "ministry_name": row.get("ministry_name"),
        "fiscal_year": year,
        "strategic_goal_number": row.get("strategic_goal_number"),
        "program_goal_number": row.get("program_goal_number"),
        "source_program_code": row.get("source_program_code"),
        "performance_program_name": row.get("performance_program_name"),
        "manual_indicator_name_plan": plan_name,
        "manual_indicator_name_report": report_name,
        "manual_indicator_unit": row.get("indicator_unit"),
        "manual_indicator_direction_raw": direction_raw,
        "manual_planned_target_raw": row.get("planned_target_raw"),
        "manual_actual_value_raw": row.get("actual_value_raw"),
        "manual_official_achievement_rate_raw": row.get("official_achievement_rate_raw"),
        "pdf_plan_program_name": None,
        "pdf_plan_indicator_name": pdf_plan_indicator_name,
        "pdf_plan_unit": None,
        "pdf_plan_direction_raw": None,
        "pdf_plan_target_raw": pdf_plan_target_raw,
        "documented_change_target_before_raw": documented_change_target_before_raw,
        "documented_change_target_after_raw": documented_change_target_after_raw,
        "documented_change_reason_raw": documented_change_reason_raw,
        "documented_change_source_file": documented_change_source_file,
        "documented_change_split_pdf_page": documented_change_split_pdf_page,
        "documented_change_source_pdf_page": documented_change_source_pdf_page,
        "documented_change_printed_page": documented_change_printed_page,
        "documented_change_ocr_conflict": documented_change_ocr_conflict,
        "documented_change_report_conflict": documented_change_report_conflict,
        "pdf_report_program_name": None,
        "pdf_report_indicator_name": pdf_report_indicator_name,
        "pdf_report_unit": None,
        "pdf_report_target_raw": pdf_report_target_raw,
        "pdf_report_actual_raw": pdf_report_actual_raw,
        "pdf_report_official_achievement_rate_raw": pdf_report_rate_raw,
        "planned_target_numeric_manual": manual_target,
        "planned_target_numeric_pdf": planned_target_numeric_pdf,
        "report_target_numeric_pdf": report_target_numeric_pdf,
        "actual_value_numeric_manual": manual_actual,
        "actual_value_numeric_pdf": actual_value_numeric_pdf,
        "official_achievement_rate_numeric_manual": manual_rate,
        "official_achievement_rate_numeric_pdf": official_rate_numeric_pdf,
        "computed_achievement_rate": computed_achievement_rate,
        "plan_target_change_abs": plan_target_change_abs,
        "plan_target_change_pct": plan_target_change_pct,
        "report_target_change_abs": report_target_change_abs,
        "report_target_change_pct": report_target_change_pct,
        "report_actual_change_abs": report_actual_change_abs,
        "report_actual_change_pct": report_actual_change_pct,
        "report_achievement_rate_change_abs_pp": report_achievement_rate_change_abs_pp,
        "plan_vs_report_target_change_abs": plan_vs_report_target_change_abs,
        "plan_vs_report_target_change_pct": plan_vs_report_target_change_pct,
        "plan_name_match_status": plan_name_match_status,
        "plan_target_match_status": plan_target_match_status,
        "report_name_match_status": report_name_match_status,
        "report_target_match_status": report_target_match_status,
        "report_actual_match_status": report_actual_match_status,
        "report_achievement_rate_match_status": report_achievement_rate_match_status,
        "page_evidence_status": page_evidence_status,
        "ocr_status": ocr_status,
        "overall_reconciliation_status": overall_status,
        "review_reason": ";".join(review_reasons) if review_reasons else None,
        # review_instruction은 build_reconciliation_table에서 이 행이 최종
        # overall_reconciliation_status로 확정된 뒤(AMBIGUOUS 재분류 포함)
        # 일괄 계산해 채웁니다. 여기서는 자리표시자로 None을 둡니다.
        "review_instruction": None,
        "reviewer": None,
        "review_status": None,
        "review_note": None,
        "review_confirmed_at": None,
        "plan_source_file": plan_spec.filename,
        "plan_split_pdf_page": plan_split_page,
        "plan_source_pdf_page": plan_source_page,
        "plan_printed_page": plan_printed_page,
        "plan_source_text": plan_source_text,
        "report_source_file": report_source_file,
        "report_split_pdf_page": report_split_page,
        "report_source_pdf_page": report_source_page,
        "report_printed_page": report_printed_page,
        "report_source_text": report_source_text,
        "source_trace": (
            f"plan:{plan_target_source};plan_method:{plan_extraction_method};"
            f"report_method:{report_extraction_method}"
        ),
    }


def _flag_indicator_name_collisions(
    manual_df: pd.DataFrame, result_df: pd.DataFrame
) -> pd.DataFrame:
    """같은 연도에 완전히 동일한 지표명을 쓰는 프로그램이 있으면 `AMBIGUOUS`로
    재분류합니다.

    `extract_plan_evidence`/`extract_report_achievement_evidence`/
    `find_change_evidence`는 모두 "그 연도 PDF 안에서 이 지표명과 일치하는
    첫 번째 위치"만 찾습니다. 같은 연도의 서로 다른 프로그램이 우연히 같은
    지표명을 쓰면(예: 2022년 "자금공급 수혜 중소기업 매출액 증가율"이
    Ⅱ-1과 Ⅲ-1에 모두 있음), 두 수기 행이 같은 PDF 근거 딱 하나를 그대로
    나눠 갖게 되어 한쪽(또는 둘 다) 값이 실제로는 다른 프로그램의 값일
    위험이 있습니다. 값을 추정해서 고치지 않고, 사람이 원문에서 프로그램별
    위치를 직접 확인해야 한다는 뜻으로 `AMBIGUOUS`를 부여합니다. 이미
    "PDF에서 못 찾음"으로 판정된 쪽은(잘못 가져올 근거 자체가 없으므로)
    건드리지 않습니다.
    """
    result_df = result_df.copy()
    for year, sub in manual_df.groupby("fiscal_year"):
        plan_counts = sub["indicator_name_plan"].dropna().map(normalize_indicator_name)
        plan_counts = plan_counts[plan_counts != ""].value_counts()
        report_counts = sub["indicator_name_report"].dropna().map(normalize_indicator_name)
        report_counts = report_counts[report_counts != ""].value_counts()
        dup_plan_names = set(plan_counts[plan_counts > 1].index)
        dup_report_names = set(report_counts[report_counts > 1].index)
        if not dup_plan_names and not dup_report_names:
            continue

        year_idx = result_df.index[result_df["fiscal_year"] == int(year)]
        for idx in year_idx:
            plan_norm = normalize_indicator_name(result_df.at[idx, "manual_indicator_name_plan"])
            report_norm = normalize_indicator_name(
                result_df.at[idx, "manual_indicator_name_report"]
            )
            plan_dup = plan_norm in dup_plan_names
            report_dup = report_norm in dup_report_names
            if not plan_dup and not report_dup:
                continue

            if plan_dup and result_df.at[idx, "plan_name_match_status"] != "PDF_NOT_FOUND":
                result_df.at[idx, "plan_name_match_status"] = "AMBIGUOUS"
            if report_dup and result_df.at[idx, "report_name_match_status"] not in (
                "PDF_NOT_FOUND",
                "OCR_REQUIRED",
            ):
                result_df.at[idx, "report_name_match_status"] = "AMBIGUOUS"

            existing_reasons = [
                r for r in (result_df.at[idx, "review_reason"] or "").split(";") if r
            ]
            existing_reasons.append("INDICATOR_NAME_AMBIGUOUS_MULTIPLE_PROGRAMS_SAME_YEAR")
            result_df.at[idx, "review_reason"] = ";".join(existing_reasons)

            result_df.at[idx, "overall_reconciliation_status"] = _pick_overall_status(
                [
                    result_df.at[idx, "plan_name_match_status"],
                    result_df.at[idx, "plan_target_match_status"],
                    result_df.at[idx, "report_name_match_status"],
                    result_df.at[idx, "report_target_match_status"],
                    result_df.at[idx, "report_actual_match_status"],
                    result_df.at[idx, "report_achievement_rate_match_status"],
                    result_df.at[idx, "ocr_status"],
                ]
            )
    return result_df


def _fmt_page(value: Any) -> str:
    """페이지 번호를 사람이 읽기 좋은 정수 문자열로 만듭니다. 결측이면 물음표."""
    if value is None:
        return "?"
    if isinstance(value, float) and pd.isna(value):
        return "?"
    try:
        return str(int(value))
    except (TypeError, ValueError):
        return str(value)


def _build_review_instruction(row: Mapping[str, Any]) -> str | None:
    """ "어느 파일 몇 쪽을 보라"를 한 줄로 모읍니다.

    `review_reason`은 문제 유형(코드)만 담고 있어 사람이 매번 plan_source_file·
    plan_split_pdf_page 등 여러 컬럼을 직접 조합해야 했습니다. 이 함수는 새 값을
    추정하지 않고, 이미 계산된 plan/report/documented_change 파일·쪽번호 컬럼만
    골라 붙여서 검토자가 바로 파일을 열어볼 수 있게 합니다.
    """
    status = row.get("overall_reconciliation_status")
    if status in (None, "EXACT_MATCH", "NOT_APPLICABLE"):
        return None

    ok_statuses = {"EXACT_MATCH", "NOT_APPLICABLE", None}
    plan_issue = (
        row.get("plan_name_match_status") not in ok_statuses
        or row.get("plan_target_match_status") not in ok_statuses
    )
    report_issue = (
        row.get("report_name_match_status") not in ok_statuses
        or row.get("report_target_match_status") not in ok_statuses
        or row.get("report_actual_match_status") not in ok_statuses
        or row.get("report_achievement_rate_match_status") not in ok_statuses
    )
    before_raw = row.get("documented_change_target_before_raw")
    after_raw = row.get("documented_change_target_after_raw")
    has_change_table = (before_raw not in (None, "")) or (after_raw not in (None, ""))

    parts: list[str] = []
    review_reason = str(row.get("review_reason") or "")
    if status == "AMBIGUOUS" and "INDICATOR_NAME_AMBIGUOUS" in review_reason:
        parts.append("동일 연도 내 지표명 중복 - 프로그램별 위치를 직접 대조하세요.")

    if plan_issue and row.get("plan_source_file"):
        parts.append(
            f"[계획서] {row['plan_source_file']} {_fmt_page(row.get('plan_split_pdf_page'))}쪽"
            f"(원본 {_fmt_page(row.get('plan_source_pdf_page'))}쪽)"
        )
    if report_issue and row.get("report_source_file"):
        parts.append(
            f"[보고서] {row['report_source_file']} "
            f"{_fmt_page(row.get('report_split_pdf_page'))}쪽"
            f"(원본 {_fmt_page(row.get('report_source_pdf_page'))}쪽)"
        )
    if has_change_table and row.get("documented_change_split_pdf_page") is not None:
        change_file = row.get("documented_change_source_file")
        parts.append(
            f"[별첨6 변경표]{f' {change_file}' if change_file else ''} "
            f"{_fmt_page(row.get('documented_change_split_pdf_page'))}쪽"
            f"(원본 {_fmt_page(row.get('documented_change_source_pdf_page'))}쪽)"
        )
    if not parts:
        # PDF_NOT_FOUND처럼 볼 페이지 자체가 없는 경우 review_reason을 그대로 노출합니다.
        return row.get("review_reason")
    return " / ".join(parts)


MANUAL_REVIEW_CONFIRMATIONS_COLUMNS: tuple[str, ...] = (
    "source_indicator_id",
    "reviewer",
    "review_status",
    "review_note",
    "review_confirmed_at",
)

DEFAULT_MANUAL_REVIEW_CONFIRMATIONS_PATH = Path(
    "data/manual/performance/pdf_reconciliation_manual_confirmations.csv"
)


def load_manual_review_confirmations(path: Path) -> pd.DataFrame:
    """사람이 원문을 직접 확인한 뒤 채운 검수 확정 파일을 읽습니다.

    이 파일은 파이프라인이 자동 생성하지 않습니다. 사람(또는 사람을 대신해
    화면을 같이 보고 확인받은 에이전트)이 원본 PDF·이미지를 직접 봐서 값이
    맞는지 확인한 뒤에만 행을 추가하는, 검수 결과를 담는 별도 입력 파일입니다.
    파일이 없으면 아직 확정된 검수가 없다는 뜻으로 빈 DataFrame을 돌려줍니다.
    """
    if not path.is_file():
        return pd.DataFrame(columns=MANUAL_REVIEW_CONFIRMATIONS_COLUMNS)
    df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    missing = [c for c in MANUAL_REVIEW_CONFIRMATIONS_COLUMNS if c not in df.columns]
    if missing:
        raise PdfReconciliationError(f"검수 확정 파일에 필수 컬럼이 없습니다: {missing}")
    bad_status = sorted({s for s in df["review_status"] if s not in REVIEW_STATUS_VALUES})
    if bad_status:
        raise PdfReconciliationError(
            f"검수 확정 파일의 review_status에 허용되지 않는 값이 있습니다: {bad_status}"
        )
    dup = df.loc[df["source_indicator_id"].duplicated(), "source_indicator_id"].tolist()
    if dup:
        raise PdfReconciliationError(f"검수 확정 파일에 중복 source_indicator_id가 있습니다: {dup}")
    return df


def apply_manual_review_confirmations(
    result_df: pd.DataFrame, confirmations_df: pd.DataFrame
) -> pd.DataFrame:
    """검수 확정 파일 내용을 reviewer/review_status/review_note/review_confirmed_at에 반영합니다.

    - 확정 파일에 없는 행은 손대지 않습니다(아직 미확정이라는 뜻으로 None 유지).
    - 확정 파일에 result_df에 없는 source_indicator_id가 있으면 조용히 무시하지
      않고 즉시 오류를 냅니다(오탈자·행 삭제로 인한 검수 유실을 막기 위함).
    - overall_reconciliation_status나 수치값은 절대 바꾸지 않습니다. 검수는
      "확인했다"는 사람의 기록일 뿐, 원본 데이터를 대체하지 않습니다.
    """
    result_df = result_df.copy()
    if confirmations_df.empty:
        return result_df
    unknown_ids = sorted(
        set(confirmations_df["source_indicator_id"]) - set(result_df["source_indicator_id"])
    )
    if unknown_ids:
        raise PdfReconciliationError(
            f"검수 확정 파일에 result_df에 없는 source_indicator_id가 있습니다: {unknown_ids}"
        )
    confirmations_indexed = confirmations_df.set_index("source_indicator_id")
    for col in ("reviewer", "review_status", "review_note", "review_confirmed_at"):
        mapped = result_df["source_indicator_id"].map(confirmations_indexed[col].to_dict())
        result_df[col] = mapped.where(mapped.notna(), result_df.get(col))
    return result_df


def build_reconciliation_table(
    manual_df: pd.DataFrame,
    *,
    doc_specs: tuple[PdfDocSpec, ...] = PDF_DOC_SPECS,
    plan_image_only_by_year: dict[int, bool] | None = None,
    plan_max_page_count_by_year: dict[int, int] | None = None,
) -> pd.DataFrame:
    """수기 성과지표 전체를 지정한 PDF 근거와 대조합니다."""
    missing_cols = [c for c in REQUIRED_MANUAL_COLUMNS if c not in manual_df.columns]
    if missing_cols:
        raise PdfReconciliationError(f"수기 데이터에 필수 컬럼이 없습니다: {missing_cols}")

    years = sorted(manual_df["fiscal_year"].dropna().unique().tolist())
    plan_evidence_by_year: dict[int, dict[str, PlanEvidence]] = {}
    report_evidence_by_year: dict[int, dict[str, AchievementEvidence]] = {}
    plan_valid_pages: dict[int, tuple[int, int]] = {}
    report_valid_pages: dict[int, tuple[int, int]] = {}
    report_has_pua_by_year: dict[int, bool] = {}
    report_structure_review_by_year: dict[int, bool] = {}

    for year in years:
        year = int(year)
        sub = manual_df[manual_df["fiscal_year"] == year]
        plan_spec = doc_spec(year, "plan", doc_specs)
        report_spec = doc_spec(year, "report", doc_specs)
        plan_evidence_by_year[year] = extract_plan_evidence(
            plan_spec,
            sub["indicator_name_plan"].dropna().tolist(),
            image_only=(plan_image_only_by_year or PLAN_TABLE_IS_IMAGE_ONLY).get(year, True),
            max_page_count=(plan_max_page_count_by_year or {}).get(year),
        )
        report_evidence_by_year[year] = extract_report_achievement_evidence(
            report_spec, sub["indicator_name_report"].dropna().tolist()
        )
        report_names = sub["indicator_name_report"].dropna().tolist()
        full_report_text = normalize_indicator_name("".join(load_page_texts(report_spec.path)))
        report_structure_review_by_year[year] = not report_evidence_by_year[year] and any(
            normalize_indicator_name(name) in full_report_text for name in report_names
        )
        plan_valid_pages[year] = (plan_spec.source_page_start, plan_spec.source_page_end)
        full_report_path = full_document_path(report_spec)
        if full_report_path is None:
            report_valid_pages[year] = (
                report_spec.source_page_start,
                report_spec.source_page_end,
            )
        else:
            with fitz.open(full_report_path) as full_report:
                report_valid_pages[year] = (1, len(full_report))
        report_has_pua_by_year[year] = report_section_has_pua_glyphs(report_spec)

    rows = [
        reconcile_row(
            row,
            plan_evidence_by_year=plan_evidence_by_year,
            report_evidence_by_year=report_evidence_by_year,
            plan_specs_valid_pages=plan_valid_pages,
            report_specs_valid_pages=report_valid_pages,
            report_has_pua_by_year=report_has_pua_by_year,
            report_structure_review_by_year=report_structure_review_by_year,
            plan_image_only_by_year=plan_image_only_by_year,
            doc_specs=doc_specs,
        )
        for row in manual_df.to_dict("records")
    ]
    result_df = pd.DataFrame(rows)
    result_df = _flag_indicator_name_collisions(manual_df, result_df)
    # AMBIGUOUS 재분류까지 끝난 뒤의 overall_reconciliation_status를 기준으로
    # "어느 파일 몇 쪽을 보라"는 안내문을 만듭니다(먼저 만들면 재분류 전 상태로
    # 안내문이 굳어버립니다).
    result_df["review_instruction"] = [
        _build_review_instruction(r) for r in result_df.to_dict("records")
    ]
    return result_df


# ---------------------------------------------------------------------------
# 산출물 작성과 검증 요약
# ---------------------------------------------------------------------------

FINAL_SCHEMA_COLUMNS: tuple[str, ...] = (
    "source_indicator_id",
    "ministry_code",
    "ministry_name",
    "fiscal_year",
    "strategic_goal_number",
    "program_goal_number",
    "source_program_code",
    "performance_program_name",
    "manual_indicator_name_plan",
    "manual_indicator_name_report",
    "manual_indicator_unit",
    "manual_indicator_direction_raw",
    "manual_planned_target_raw",
    "manual_actual_value_raw",
    "manual_official_achievement_rate_raw",
    "pdf_plan_program_name",
    "pdf_plan_indicator_name",
    "pdf_plan_unit",
    "pdf_plan_direction_raw",
    "pdf_plan_target_raw",
    "documented_change_target_before_raw",
    "documented_change_target_after_raw",
    "documented_change_reason_raw",
    "documented_change_source_file",
    "documented_change_split_pdf_page",
    "documented_change_source_pdf_page",
    "documented_change_printed_page",
    "documented_change_ocr_conflict",
    "documented_change_report_conflict",
    "pdf_report_program_name",
    "pdf_report_indicator_name",
    "pdf_report_unit",
    "pdf_report_target_raw",
    "pdf_report_actual_raw",
    "pdf_report_official_achievement_rate_raw",
    "planned_target_numeric_manual",
    "planned_target_numeric_pdf",
    "report_target_numeric_pdf",
    "actual_value_numeric_manual",
    "actual_value_numeric_pdf",
    "official_achievement_rate_numeric_manual",
    "official_achievement_rate_numeric_pdf",
    "computed_achievement_rate",
    "plan_target_change_abs",
    "plan_target_change_pct",
    "report_target_change_abs",
    "report_target_change_pct",
    "report_actual_change_abs",
    "report_actual_change_pct",
    "report_achievement_rate_change_abs_pp",
    "plan_vs_report_target_change_abs",
    "plan_vs_report_target_change_pct",
    "plan_name_match_status",
    "plan_target_match_status",
    "report_name_match_status",
    "report_target_match_status",
    "report_actual_match_status",
    "report_achievement_rate_match_status",
    "page_evidence_status",
    "ocr_status",
    "overall_reconciliation_status",
    "review_reason",
    "review_instruction",
    "reviewer",
    "review_status",
    "review_note",
    "review_confirmed_at",
    "plan_source_file",
    "plan_split_pdf_page",
    "plan_source_pdf_page",
    "plan_printed_page",
    "plan_source_text",
    "report_source_file",
    "report_split_pdf_page",
    "report_source_pdf_page",
    "report_printed_page",
    "report_source_text",
    "source_trace",
)


def _count_by(series: pd.Series) -> dict[str, int]:
    return {str(k): int(v) for k, v in series.value_counts(dropna=False).items()}


def build_reconciliation_summary(
    result_df: pd.DataFrame,
    *,
    manual_input_rows: int,
    source_hashes_before: dict[str, str],
    source_hashes_after: dict[str, str],
) -> dict[str, Any]:
    """10.3절 검증 요약(JSON)에 필요한 항목을 계산합니다."""
    n = len(result_df)
    hash_mismatches = {
        path: {"before": source_hashes_before.get(path), "after": source_hashes_after.get(path)}
        for path in source_hashes_before
        if source_hashes_before.get(path) != source_hashes_after.get(path)
    }
    return {
        "generated_at": datetime.now(UTC).isoformat(),
        "input_row_count": manual_input_rows,
        "output_row_count": n,
        "rows_by_fiscal_year": _count_by(result_df["fiscal_year"]),
        "rows_by_overall_status": _count_by(result_df["overall_reconciliation_status"]),
        "plan_name_match_counts": _count_by(result_df["plan_name_match_status"]),
        "report_name_match_counts": _count_by(result_df["report_name_match_status"]),
        "plan_target_match_counts": _count_by(result_df["plan_target_match_status"]),
        "report_actual_match_counts": _count_by(result_df["report_actual_match_status"]),
        "report_achievement_rate_match_counts": _count_by(
            result_df["report_achievement_rate_match_status"]
        ),
        "ocr_required_row_count": int((result_df["ocr_status"] == "OCR_REQUIRED").sum()),
        "ocr_required_pages": sorted(
            {
                p
                for p in pd.concat(
                    [
                        result_df.loc[
                            result_df["ocr_status"] == "OCR_REQUIRED", "plan_split_pdf_page"
                        ],
                        result_df.loc[
                            result_df["ocr_status"] == "OCR_REQUIRED", "report_split_pdf_page"
                        ],
                    ]
                )
                .dropna()
                .tolist()
            }
        ),
        "ambiguous_row_count": int(
            (result_df["overall_reconciliation_status"] == "AMBIGUOUS").sum()
        ),
        "pdf_not_found_row_count": int(
            (result_df["overall_reconciliation_status"] == "PDF_NOT_FOUND").sum()
        ),
        "manual_missing_pdf_present_count": int(
            (
                (result_df["plan_target_match_status"] == "MANUAL_MISSING_PDF_PRESENT")
                | (result_df["report_target_match_status"] == "MANUAL_MISSING_PDF_PRESENT")
                | (result_df["report_actual_match_status"] == "MANUAL_MISSING_PDF_PRESENT")
                | (
                    result_df["report_achievement_rate_match_status"]
                    == "MANUAL_MISSING_PDF_PRESENT"
                )
            ).sum()
        ),
        "pdf_missing_manual_present_count": int(
            (
                (result_df["plan_target_match_status"] == "PDF_MISSING_MANUAL_PRESENT")
                | (result_df["report_target_match_status"] == "PDF_MISSING_MANUAL_PRESENT")
                | (result_df["report_actual_match_status"] == "PDF_MISSING_MANUAL_PRESENT")
                | (
                    result_df["report_achievement_rate_match_status"]
                    == "PDF_MISSING_MANUAL_PRESENT"
                )
            ).sum()
        ),
        "primary_key_duplicate_count": int(result_df["source_indicator_id"].duplicated().sum()),
        "manual_review_csv_row_count": int(
            (result_df["overall_reconciliation_status"] != "EXACT_MATCH").sum()
        ),
        "review_status_counts": _count_by(result_df["review_status"]),
        "plan_vs_report_target_change": _target_change_summary(result_df),
        "source_file_sha256": source_hashes_after,
        "source_file_hash_unchanged": len(hash_mismatches) == 0,
        "source_file_hash_mismatches": hash_mismatches,
    }


def _target_change_summary(result_df: pd.DataFrame) -> dict[str, Any]:
    """계획서-보고서 목표치 변경 폭을 "얼마나" 바뀌었는지 기준으로 요약합니다.

    변경 사유는 판단하지 않고, 절대·상대 변화량 분포와 상위 변경 행만
    보여줍니다. 10%p/50%p 구간은 절대 기준이 아니라 스캔을 돕는 서술적
    구간이며, 정책적 임계값이 아닙니다.
    """
    all_sub = result_df[result_df["plan_vs_report_target_change_abs"].notna()].copy()
    unverified_ocr = all_sub[all_sub["ocr_status"].eq("OCR_REQUIRED")]
    sub = all_sub[~all_sub["ocr_status"].eq("OCR_REQUIRED")].copy()
    if sub.empty:
        return {
            "rows_with_both_pdf_targets": 0,
            "rows_excluded_unverified_ocr": len(unverified_ocr),
            "rows_unchanged": 0,
            "rows_changed": 0,
            "abs_pct_over_10": 0,
            "abs_pct_over_50": 0,
            "median_abs_pct_when_changed": None,
            "top_changes": [],
        }
    sub["abs_change"] = sub["plan_vs_report_target_change_abs"].abs()
    changed = sub[sub["abs_change"] > 1e-9]
    with_pct = changed[changed["plan_vs_report_target_change_pct"].notna()].copy()
    with_pct["abs_pct"] = with_pct["plan_vs_report_target_change_pct"].abs()
    top = with_pct.sort_values("abs_pct", ascending=False).head(10)
    return {
        "rows_with_both_pdf_targets": len(sub),
        "rows_excluded_unverified_ocr": len(unverified_ocr),
        "rows_unchanged": int(len(sub) - len(changed)),
        "rows_changed": len(changed),
        "abs_pct_over_10": int((with_pct["abs_pct"] >= 10).sum()),
        "abs_pct_over_50": int((with_pct["abs_pct"] >= 50).sum()),
        "median_abs_pct_when_changed": (
            float(with_pct["abs_pct"].median()) if not with_pct.empty else None
        ),
        "top_changes": [
            {
                "source_indicator_id": r["source_indicator_id"],
                "plan_target_pdf": r["planned_target_numeric_pdf"],
                "report_target_pdf": r["report_target_numeric_pdf"],
                "change_abs": r["plan_vs_report_target_change_abs"],
                "change_pct": r["plan_vs_report_target_change_pct"],
                # ocr_status가 OCR_REQUIRED면 계획서 쪽 값이 로컬 OCR 추출값이라
                # 변경폭 자체가 실제 목표 조정이 아니라 OCR 오인식일 수 있습니다.
                # 이 경우는 변경 크기를 확정값으로 읽지 말고 원문 이미지를
                # 먼저 확인해야 합니다.
                "ocr_status": r.get("ocr_status"),
            }
            for _, r in top.iterrows()
        ],
    }


def write_reconciliation_outputs(
    result_df: pd.DataFrame,
    *,
    manual_input_rows: int,
    source_hashes_before: dict[str, str],
    source_hashes_after: dict[str, str],
    output_dir: Path,
    output_stem: str = "mss_performance",
    overwrite: bool = False,
) -> dict[str, Path]:
    """10.1~10.3절 산출물(Parquet, 검토 CSV, 요약 JSON)을 씁니다."""
    output_dir.mkdir(parents=True, exist_ok=True)
    ordered = result_df.reindex(columns=list(FINAL_SCHEMA_COLUMNS))

    parquet_path = output_dir / f"{output_stem}_pdf_reconciliation.parquet"
    csv_path = output_dir / f"{output_stem}_pdf_manual_review.csv"
    summary_path = output_dir / "reconciliation_summary.json"

    for path in (parquet_path, csv_path, summary_path):
        if path.exists() and not overwrite:
            raise FileExistsError(f"이미 존재합니다 (overwrite=False): {path}")

    ordered.to_parquet(parquet_path, index=False)

    review_df = ordered[ordered["overall_reconciliation_status"] != "EXACT_MATCH"].copy()
    review_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    summary = build_reconciliation_summary(
        ordered,
        manual_input_rows=manual_input_rows,
        source_hashes_before=source_hashes_before,
        source_hashes_after=source_hashes_after,
    )
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"parquet": parquet_path, "manual_review_csv": csv_path, "summary_json": summary_path}


KNOWN_PLAN_IMAGE_ONLY_YEARS: dict[str, set[int]] = {
    "019": {2022},
    "075": {2022, 2024},
    "162": {2022},
}
KNOWN_PLAN_MAX_PAGE_COUNTS: dict[tuple[str, int], int] = {
    # 2024 복지부 계획서는 텍스트 레이어가 깨져 별첨2 문자열 탐색이 불가능합니다.
    # 렌더링 전수 점검 결과 별첨1은 분리 PDF 1~55쪽, 별첨2는 56쪽부터입니다.
    ("075", 2024): 55,
}


def run_ministry_pdf_reconciliation(
    ministry_code: str,
    *,
    manual_parquet_path: Path | None = None,
    manual_excel_path: Path = Path("data/manual/LLM_문서구조화_3개부처_최종제출본.xlsx"),
    output_root: Path = Path("data/processed/performance/pdf_reconciliation"),
    overwrite: bool = False,
) -> dict[str, Any]:
    """한 부처의 2022~2024 수기 성과지표를 분리 PDF 원문과 대조합니다."""
    ministry_code = str(ministry_code).zfill(3)
    if ministry_code not in KNOWN_PLAN_IMAGE_ONLY_YEARS:
        raise PdfReconciliationError(
            f"계획서 렌더링 상태를 검증하지 않은 부처입니다: {ministry_code}"
        )
    manual_parquet_path = manual_parquet_path or Path(
        f"data/processed/performance/by_ministry/ministry_code={ministry_code}/"
        "program_kpi_year.parquet"
    )
    if not manual_parquet_path.is_file():
        raise PdfReconciliationError(f"기준 파일을 찾을 수 없습니다: {manual_parquet_path}")
    if not manual_excel_path.is_file():
        raise PdfReconciliationError(f"수기 원본을 찾을 수 없습니다: {manual_excel_path}")

    manual_df = pd.read_parquet(manual_parquet_path)
    if manual_df.empty:
        raise PdfReconciliationError(f"기준 파일이 비어 있습니다: {manual_parquet_path}")
    doc_specs = discover_pdf_doc_specs(ministry_code)
    image_only = {
        year: year in KNOWN_PLAN_IMAGE_ONLY_YEARS[ministry_code] for year in (2022, 2023, 2024)
    }
    max_pages = {
        year: count
        for (code, year), count in KNOWN_PLAN_MAX_PAGE_COUNTS.items()
        if code == ministry_code
    }

    hashes_before = all_source_hashes(manual_excel_path, doc_specs)
    hashes_before[str(manual_parquet_path)] = sha256_file(manual_parquet_path)
    result_df = build_reconciliation_table(
        manual_df,
        doc_specs=doc_specs,
        plan_image_only_by_year=image_only,
        plan_max_page_count_by_year=max_pages,
    )
    hashes_after = all_source_hashes(manual_excel_path, doc_specs)
    hashes_after[str(manual_parquet_path)] = sha256_file(manual_parquet_path)

    output_dir = output_root / f"ministry_code={ministry_code}"
    output_paths = write_reconciliation_outputs(
        result_df,
        manual_input_rows=len(manual_df),
        source_hashes_before=hashes_before,
        source_hashes_after=hashes_after,
        output_dir=output_dir,
        output_stem=f"{ministry_code}_performance",
        overwrite=overwrite,
    )
    inventory_path = output_dir / "pdf_page_inventory.csv"
    if inventory_path.exists() and not overwrite:
        raise FileExistsError(f"이미 존재합니다 (overwrite=False): {inventory_path}")
    build_page_inventory(run_ocr=False, doc_specs=doc_specs).to_csv(
        inventory_path, index=False, encoding="utf-8-sig"
    )
    output_paths["page_inventory_csv"] = inventory_path
    summary = build_reconciliation_summary(
        result_df,
        manual_input_rows=len(manual_df),
        source_hashes_before=hashes_before,
        source_hashes_after=hashes_after,
    )
    return {"output_paths": output_paths, "summary": summary, "result_df": result_df}


def run_pdf_reconciliation(
    *,
    manual_parquet_path: Path = Path("data/processed/performance/program_kpi_year.parquet"),
    manual_excel_path: Path = Path("data/manual/LLM_문서구조화_중기부_최종.xlsx"),
    output_dir: Path = Path("data/processed/performance/pdf_reconciliation"),
    export_dir: Path = Path("data/exports/performance"),
    manual_review_confirmations_path: Path = DEFAULT_MANUAL_REVIEW_CONFIRMATIONS_PATH,
    overwrite: bool = False,
) -> dict[str, Any]:
    """63행 전체 대조 파이프라인을 실행하고 산출물·요약을 반환합니다."""
    if not manual_parquet_path.is_file():
        raise PdfReconciliationError(f"기준 파일을 찾을 수 없습니다: {manual_parquet_path}")

    manual_df = pd.read_parquet(manual_parquet_path)
    input_row_count = len(manual_df)
    if input_row_count != 63:
        raise PdfReconciliationError(
            f"기준 파일이 63행이 아닙니다 (실제 {input_row_count}행): {manual_parquet_path}"
        )

    hashes_before = all_source_hashes(manual_excel_path)
    hashes_before[str(manual_parquet_path)] = sha256_file(manual_parquet_path)

    result_df = build_reconciliation_table(manual_df)
    confirmations_df = load_manual_review_confirmations(manual_review_confirmations_path)
    result_df = apply_manual_review_confirmations(result_df, confirmations_df)

    hashes_after = all_source_hashes(manual_excel_path)
    hashes_after[str(manual_parquet_path)] = sha256_file(manual_parquet_path)

    output_paths = write_reconciliation_outputs(
        result_df,
        manual_input_rows=input_row_count,
        source_hashes_before=hashes_before,
        source_hashes_after=hashes_after,
        output_dir=output_dir,
        overwrite=overwrite,
    )
    summary = build_reconciliation_summary(
        result_df,
        manual_input_rows=input_row_count,
        source_hashes_before=hashes_before,
        source_hashes_after=hashes_after,
    )

    page_inventory = build_page_inventory(run_ocr=False)
    excel_path = export_dir / "mss_performance_pdf_reconciliation.xlsx"
    if excel_path.exists() and not overwrite:
        raise FileExistsError(f"이미 존재합니다 (overwrite=False): {excel_path}")
    write_reconciliation_excel(result_df, summary, page_inventory, excel_path, overwrite=overwrite)
    output_paths["excel"] = excel_path

    return {"output_paths": output_paths, "summary": summary, "result_df": result_df}


# ---------------------------------------------------------------------------
# 10.4절 사람 검토용 엑셀
# ---------------------------------------------------------------------------

NAVY = "17365D"
PALE_BLUE = "EAF3F8"
GREEN = "E2F0D9"
BLUE = "D9EAF7"
YELLOW = "FFF2CC"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"

STATUS_COLOR = {
    "EXACT_MATCH": GREEN,
    "MATCH_AFTER_CHANGE": BLUE,
    "ROUNDING_ONLY": BLUE,
    "OCR_REQUIRED": YELLOW,
    "MANUAL_REVIEW": YELLOW,
    "VALUE_MISMATCH": RED,
    "AMBIGUOUS": RED,
    "PDF_NOT_FOUND": RED,
    "NOT_APPLICABLE": GRAY,
}

REVIEW_STATUS_VALUES = ("", "PENDING", "CONFIRMED", "CORRECTED", "NOT_RESOLVABLE")


def _set_banner(sheet: Any, title: str, subtitle: str, end_column: int) -> None:
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end_letter}1")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(color=WHITE, bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.merge_cells(f"A2:{end_letter}2")
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet["A2"].font = Font(color=NAVY, italic=True)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32
    sheet.sheet_view.showGridLines = False


def _write_readme_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "README"
    _set_banner(
        sheet,
        "중기부 성과지표 PDF 대조 결과 (63행)",
        "성과계획서·성과보고서 별첨과 수기 구조화 63행을 원문 대조한 결과입니다. 아래 항목을 먼저 읽어주세요.",
        2,
    )
    lines = [
        ("목적", "수기 63행(부처×프로그램×성과지표×회계연도)이 PDF 원문과 일치하는지 확인"),
        ("입력", "program_kpi_year.parquet(63행), LLM_문서구조화_중기부_최종.xlsx(원본, 읽기전용)"),
        (
            "PDF 근거",
            "계획서 별첨1(목표치, 2022 OCR), 보고서 별첨3 3.세부현황(목표·실적·달성률), 보고서 별첨6(변경사항)",
        ),
        ("한 행의 기준", "부처 × 프로그램 × 성과지표 × 회계연도 (source_indicator_id)"),
        ("EXACT_MATCH", "수기값과 PDF 원문이 허용오차 내로 일치"),
        ("MATCH_AFTER_CHANGE", "별첨6 변경사항표로 계획서-보고서 지표명·목표치 변경이 확인됨"),
        ("ROUNDING_ONLY", "달성률 차이가 0.1%p 이내(반올림 차이로 추정)"),
        ("VALUE_MISMATCH", "수기값과 PDF 원문 값이 허용오차를 넘어 다름 (원문 재확인 필요)"),
        ("MANUAL_MISSING_PDF_PRESENT", "수기 데이터는 결측이나 PDF 원문에는 값이 있음"),
        ("PDF_MISSING_MANUAL_PRESENT", "PDF 원문에서 값을 찾지 못했으나 수기 데이터에는 값이 있음"),
        ("PDF_NOT_FOUND", "해당 연도 별첨 전체에서 지표를 찾지 못함"),
        (
            "OCR_REQUIRED",
            "이미지 표(2022 별첨1) 또는 폰트 매핑 오류 페이지로 로컬 OCR을 썼거나 써야 함. 사람 확인 필요",
        ),
        ("AMBIGUOUS", "PDF 지표 하나가 수기 여러 행에 대응하거나 반대인 경우"),
        ("MANUAL_REVIEW", "지표명 변경이 확인되지 않았거나 그 외 자동 확정 불가"),
        ("숫자 비교 규칙", "쉼표·%·공백만 제거해 비교. 원문 문자열(raw)과 숫자 변환값을 모두 보존"),
        (
            "달성률 계산식",
            "상향지표=실적÷목표×100, 하향지표=목표÷실적×100. 공식 달성률은 덮어쓰지 않고 computed_achievement_rate로 별도 보존",
        ),
        (
            "해석 제한",
            "이 결과는 지표 정의·원문 검증용입니다. 성과 우열이나 사업 우수/부실 판정에 사용하지 마세요.",
        ),
        ("해석 제한 2", "프로그램 성과지표를 세부사업 성과로 귀속하지 않습니다."),
        ("생성 시각", summary.get("generated_at", "")),
    ]
    for i, (label, value) in enumerate(lines, start=4):
        sheet.cell(i, 1, label).font = Font(bold=True, color=NAVY)
        sheet.cell(i, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[i].height = 30
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 110
    sheet.freeze_panes = "A4"


def _write_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("SUMMARY")
    _set_banner(sheet, "요약", "연도별·상태별 건수 (reconciliation_summary.json과 동일한 값)", 2)
    row = 4
    sections = [
        ("연도별 행 수", summary["rows_by_fiscal_year"]),
        ("전체 상태별 행 수", summary["rows_by_overall_status"]),
        ("계획서 지표명 상태", summary["plan_name_match_counts"]),
        ("보고서 지표명 상태", summary["report_name_match_counts"]),
        ("계획서 목표치 상태", summary["plan_target_match_counts"]),
        ("보고서 실적치 상태", summary["report_actual_match_counts"]),
        ("보고서 달성률 상태", summary["report_achievement_rate_match_counts"]),
    ]
    for title, counts in sections:
        sheet.cell(row, 1, title).font = Font(bold=True, color=NAVY)
        row += 1
        for key, value in counts.items():
            sheet.cell(row, 1, key)
            sheet.cell(row, 2, value)
            row += 1
        row += 1

    change = summary.get("plan_vs_report_target_change", {})
    sheet.cell(row, 1, "계획-보고 목표치 변경 (얼마나 바뀌었는지, 사유는 판단 안 함)").font = Font(
        bold=True, color=NAVY
    )
    row += 1
    for key, label in (
        ("rows_with_both_pdf_targets", "계획·보고 목표 둘 다 있는 행"),
        ("rows_unchanged", "변경 없음(동일)"),
        ("rows_changed", "변경됨"),
        ("abs_pct_over_10", "변경률 10%p 이상"),
        ("abs_pct_over_50", "변경률 50%p 이상"),
        ("median_abs_pct_when_changed", "변경된 행의 변경률 중앙값(%)"),
    ):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, change.get(key))
        row += 1
    row += 1
    sheet.cell(
        row,
        1,
        "변경폭 상위 사례 (indicator / 계획목표 / 보고목표 / 변경량 / 변경률% / "
        "OCR상태 - OCR_REQUIRED면 계획값이 OCR 추출값이라 원문 재확인 필요)",
    ).font = Font(bold=True)
    row += 1
    for item in change.get("top_changes", []):
        sheet.cell(row, 1, item.get("source_indicator_id"))
        sheet.cell(row, 2, item.get("plan_target_pdf"))
        sheet.cell(row, 3, item.get("report_target_pdf"))
        sheet.cell(row, 4, item.get("change_abs"))
        sheet.cell(row, 5, item.get("change_pct"))
        sheet.cell(row, 6, item.get("ocr_status"))
        row += 1
    row += 1

    sheet.cell(row, 1, "입력 행 수").font = Font(bold=True)
    sheet.cell(row, 2, summary["input_row_count"])
    row += 1
    sheet.cell(row, 1, "출력 행 수").font = Font(bold=True)
    sheet.cell(row, 2, summary["output_row_count"])
    row += 1
    sheet.cell(row, 1, "기본키 중복 수").font = Font(bold=True)
    sheet.cell(row, 2, summary["primary_key_duplicate_count"])
    row += 1
    sheet.cell(row, 1, "원본 해시 변경 없음").font = Font(bold=True)
    sheet.cell(row, 2, "예" if summary["source_file_hash_unchanged"] else "아니오 - 확인 필요")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 20
    sheet.freeze_panes = "A4"


def _write_data_sheet(
    workbook: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    editable_columns: tuple[str, ...] = (),
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    headers = list(df.columns)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34

    text_columns = {
        "source_indicator_id",
        "ministry_code",
        "program_goal_number",
        "source_program_code",
        "strategic_goal_number",
    }
    for r, record in enumerate(df.to_dict("records"), start=2):
        for c, header in enumerate(headers, start=1):
            value = record.get(header)
            if value is None or (isinstance(value, float) and pd.isna(value)):
                value = ""
            cell = sheet.cell(r, c, value)
            if header in text_columns:
                cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[r].height = 30

    last_row = len(df) + 1
    last_col_letter = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    sheet.freeze_panes = "C2"
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22

    if "overall_reconciliation_status" in headers and last_row >= 2:
        status_col_letter = get_column_letter(headers.index("overall_reconciliation_status") + 1)
        data_range = f"A2:{last_col_letter}{last_row}"
        for status, color in STATUS_COLOR.items():
            sheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[f'${status_col_letter}2="{status}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )

    if "review_status" in headers and editable_columns:
        review_col = headers.index("review_status") + 1
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(v or " " for v in REVIEW_STATUS_VALUES) + '"',
            allow_blank=True,
        )
        validation.error = (
            "빈 값, PENDING, CONFIRMED, CORRECTED, NOT_RESOLVABLE 중 하나여야 합니다."
        )
        validation.errorTitle = "허용되지 않는 값"
        sheet.add_data_validation(validation)
        if last_row >= 2:
            validation.add(
                f"{get_column_letter(review_col)}2:{get_column_letter(review_col)}{last_row}"
            )


def _write_page_qa_sheet(workbook: Workbook, page_inventory: pd.DataFrame) -> None:
    _write_data_sheet(workbook, "PAGE_QA", page_inventory)


def write_reconciliation_excel(
    result_df: pd.DataFrame,
    summary: dict[str, Any],
    page_inventory: pd.DataFrame,
    output_path: Path,
    *,
    overwrite: bool = False,
) -> Path:
    """10.4절 사람 검토용 엑셀(README/SUMMARY/RECONCILIATION_63/MANUAL_REVIEW/PAGE_QA)."""
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"이미 존재합니다 (overwrite=False): {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    ordered = result_df.reindex(columns=list(FINAL_SCHEMA_COLUMNS))
    review_df = ordered[ordered["overall_reconciliation_status"] != "EXACT_MATCH"].copy()
    # 이미 검수 확정 파일로 채워진 review_note는 보존하고, 아직 미확정인 행만
    # 엑셀에서 사람이 바로 타이핑할 수 있게 빈 문자열로 둡니다.
    review_df["review_note"] = review_df["review_note"].fillna("")

    workbook = Workbook()
    _write_readme_sheet(workbook, summary)
    _write_summary_sheet(workbook, summary)
    _write_data_sheet(workbook, "RECONCILIATION_63", ordered)
    _write_data_sheet(
        workbook,
        "MANUAL_REVIEW",
        review_df,
        editable_columns=("reviewer", "review_status", "review_note"),
    )
    _write_page_qa_sheet(workbook, page_inventory)
    workbook.save(output_path)
    return output_path


__all__ = [
    "ALLOWED_STATUS_VALUES",
    "DEFAULT_MANUAL_REVIEW_CONFIRMATIONS_PATH",
    "FINAL_SCHEMA_COLUMNS",
    "KNOWN_PLAN_IMAGE_ONLY_YEARS",
    "KNOWN_PLAN_MAX_PAGE_COUNTS",
    "MANUAL_REVIEW_CONFIRMATIONS_COLUMNS",
    "PDF_DOC_SPECS",
    "REQUIRED_MANUAL_COLUMNS",
    "REVIEW_STATUS_VALUES",
    "STATUS_PRIORITY",
    "AchievementEvidence",
    "ChangeEvidence",
    "PdfDocSpec",
    "PdfReconciliationError",
    "PlanEvidence",
    "all_source_hashes",
    "apply_manual_review_confirmations",
    "build_page_inventory",
    "build_reconciliation_summary",
    "build_reconciliation_table",
    "classify_numeric_match",
    "classify_rate_match",
    "clean_direction",
    "compute_achievement_rate",
    "discover_pdf_doc_specs",
    "doc_spec",
    "extract_plan_evidence",
    "extract_report_achievement_evidence",
    "find_change_evidence",
    "load_manual_review_confirmations",
    "load_page_texts",
    "normalize_indicator_name",
    "normalize_numeric_raw",
    "numeric_change",
    "ocr_page_text",
    "parse_numeric",
    "printed_page_number",
    "reconcile_row",
    "report_section_has_pua_glyphs",
    "run_ministry_pdf_reconciliation",
    "run_pdf_reconciliation",
    "select_plan_target",
    "sha256_file",
    "write_reconciliation_excel",
    "write_reconciliation_outputs",
]
