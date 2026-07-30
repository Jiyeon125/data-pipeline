"""사람이 검수한 성과 구조화 엑셀을 프로그램-연도 재정 마스터와 연결합니다."""

from __future__ import annotations

import hashlib
import json
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

GOLD_SHEET = "02_골드셋"
HEADER_ROW = 4
HEADER_MAP = {
    "행ID": "source_indicator_id",
    "부처": "ministry_name",
    "회계연도": "fiscal_year",
    "전략목표번호": "strategic_goal_number",
    "프로그램목표번호": "program_goal_number",
    "예산프로그램코드": "source_program_code",
    "프로그램명": "performance_program_name",
    "계획서 성과지표명": "indicator_name_plan",
    "보고서 성과지표명": "indicator_name_report",
    "단위": "indicator_unit",
    "지표방향": "indicator_direction",
    "목표치": "planned_target_raw",
    "실적치": "actual_value_raw",
    "달성률": "official_achievement_rate_raw",
    "계획서 페이지": "plan_source_page",
    "보고서 페이지": "report_source_page",
    "계획서 근거": "plan_source_text",
    "보고서 근거": "report_source_text",
    "매칭상태": "plan_report_match_status_raw",
    "변경내용": "change_description",
    "검수자": "reviewer",
    "비고": "notes",
}
REQUIRED_HEADERS = tuple(HEADER_MAP)
REQUIRED_FINANCIAL_COLUMNS = {
    "fiscal_year",
    "ministry_code",
    "ministry_name",
    "field_name",
    "sector_name",
    "program_code",
    "program_name",
    "original_budget",
    "current_budget",
    "settlement_expenditure",
    "execution_rate",
    "financial_linkage_status",
    "financial_quality_level",
}
AUTO_MATCH_STATUSES = {
    "EXACT_CODE",
    "EXACT_NAME",
    "NORMALIZED_NAME",
    "EXACT_NAME_UNIQUE_FINANCIAL",
}
PROGRAM_CODE_CONFIRMATION_KEY = [
    "ministry_name",
    "fiscal_year",
    "program_goal_number",
    "program_name_normalized",
]


class ManualPerformanceError(ValueError):
    """수기 성과자료가 입력·키·보존 계약을 위반할 때 발생합니다."""


@dataclass
class ManualPerformanceResult:
    manual_rows: pd.DataFrame
    indicators: pd.DataFrame
    program_year: pd.DataFrame
    issues: pd.DataFrame
    data_dictionary: pd.DataFrame
    summary: dict[str, Any]
    output_paths: list[Path]


def _text(value: Any) -> str | None:
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _year(value: Any) -> int | None:
    text = _text(value)
    if text is None or not text.isdigit():
        return None
    return int(text)


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text or text in {"-", "신규", "종료", "집계중", "해당없음"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_program_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value) or "")
    return "".join(character for character in text if not character.isspace())


def apply_program_code_confirmations(
    indicators: pd.DataFrame,
    confirmations_path: Path | None,
) -> pd.DataFrame:
    if confirmations_path is None:
        return indicators.copy()
    if not confirmations_path.is_file():
        raise ManualPerformanceError(
            f"프로그램코드 확인표를 찾을 수 없습니다: {confirmations_path}"
        )
    confirmations = pd.read_csv(
        confirmations_path,
        dtype={
            "ministry_name": "string",
            "program_goal_number": "string",
            "performance_program_name": "string",
            "source_program_code": "string",
            "source_field_name": "string",
            "source_sector_name": "string",
        },
    )
    required = {
        "ministry_name",
        "fiscal_year",
        "program_goal_number",
        "performance_program_name",
        "source_program_code",
        "mapping_status",
    }
    missing = required - set(confirmations.columns)
    if missing:
        raise ManualPerformanceError(
            f"프로그램코드 확인표 필수 열이 없습니다: {', '.join(sorted(missing))}"
        )
    confirmations["fiscal_year"] = pd.to_numeric(
        confirmations["fiscal_year"], errors="coerce"
    ).astype("Int64")
    confirmations["program_name_normalized"] = confirmations["performance_program_name"].map(
        normalize_program_name
    )
    for column in ("source_field_name", "source_sector_name"):
        if column not in confirmations:
            confirmations[column] = pd.NA
    if confirmations.duplicated(PROGRAM_CODE_CONFIRMATION_KEY).any():
        raise ManualPerformanceError("프로그램코드 확인표의 프로그램목표 키가 중복되었습니다.")

    confirmed = confirmations.loc[
        :,
        [
            *PROGRAM_CODE_CONFIRMATION_KEY,
            "source_program_code",
            "source_field_name",
            "source_sector_name",
            "mapping_status",
        ],
    ].rename(
        columns={
            "source_program_code": "confirmed_source_program_code",
            "source_field_name": "confirmed_source_field_name",
            "source_sector_name": "confirmed_source_sector_name",
            "mapping_status": "confirmed_program_mapping_status",
        }
    )
    result = indicators.merge(
        confirmed,
        how="left",
        on=PROGRAM_CODE_CONFIRMATION_KEY,
        validate="many_to_one",
    )
    conflict = (
        result["source_program_code"].notna()
        & result["confirmed_source_program_code"].notna()
        & result["source_program_code"]
        .astype(str)
        .ne(result["confirmed_source_program_code"].astype(str))
    )
    if conflict.any():
        raise ManualPerformanceError("원본과 확인표의 프로그램코드가 충돌합니다.")
    confirmed_rows = result["confirmed_source_program_code"].notna()
    result.loc[confirmed_rows, "source_program_code"] = result.loc[
        confirmed_rows, "confirmed_source_program_code"
    ]
    result["source_field_name"] = result["confirmed_source_field_name"]
    result["source_sector_name"] = result["confirmed_source_sector_name"]
    result["program_mapping_status"] = result["confirmed_program_mapping_status"]
    return result.drop(
        columns=[
            "confirmed_source_program_code",
            "confirmed_source_field_name",
            "confirmed_source_sector_name",
            "confirmed_program_mapping_status",
        ]
    ).convert_dtypes()


def _is_example(record: dict[str, Any]) -> bool:
    review_text = " ".join(
        _text(record.get(column)) or "" for column in ("reviewer", "notes", "change_description")
    )
    return "예시" in review_text


def _source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_manual_performance_workbook(
    input_path: Path,
    *,
    ministry_name: str,
    start_year: int,
    end_year: int,
) -> pd.DataFrame:
    """골드셋의 모든 비어 있지 않은 행을 출처 행번호와 함께 읽습니다."""
    if not input_path.is_file():
        raise ManualPerformanceError(f"수기 성과 엑셀을 찾을 수 없습니다: {input_path}")
    if start_year > end_year:
        raise ManualPerformanceError("시작연도는 종료연도보다 클 수 없습니다.")

    workbook = load_workbook(input_path, read_only=True, data_only=False)
    if GOLD_SHEET not in workbook.sheetnames:
        raise ManualPerformanceError(f"필수 시트가 없습니다: {GOLD_SHEET}")
    sheet = workbook[GOLD_SHEET]
    headers = [_text(cell.value) for cell in sheet[HEADER_ROW]]
    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise ManualPerformanceError(f"골드셋 필수 열이 없습니다: {', '.join(missing_headers)}")

    positions = {header: headers.index(header) for header in REQUIRED_HEADERS}
    rows: list[dict[str, Any]] = []
    for source_row_number, cells in enumerate(
        sheet.iter_rows(min_row=HEADER_ROW + 1),
        start=HEADER_ROW + 1,
    ):
        values = [cell.value for cell in cells]
        if not any(value not in (None, "") for value in values):
            continue
        record = {target: _text(values[positions[source]]) for source, target in HEADER_MAP.items()}
        record["fiscal_year"] = _year(values[positions["회계연도"]])
        record["source_file"] = input_path.name
        record["source_sheet"] = GOLD_SHEET
        record["source_row_number"] = source_row_number
        record["source_trace"] = f"{input_path.as_posix()}#{GOLD_SHEET}!row={source_row_number}"
        record["is_example"] = _is_example(record)
        record["ministry_scope_match"] = record["ministry_name"] == ministry_name
        record["year_scope_match"] = (
            record["fiscal_year"] is not None and start_year <= record["fiscal_year"] <= end_year
        )
        record["required_fields_available"] = all(
            record.get(column) not in (None, "")
            for column in (
                "source_indicator_id",
                "ministry_name",
                "fiscal_year",
                "performance_program_name",
            )
        ) and bool(record.get("indicator_name_plan") or record.get("indicator_name_report"))
        record["analysis_eligible"] = (
            not record["is_example"]
            and record["ministry_scope_match"]
            and record["year_scope_match"]
            and record["required_fields_available"]
        )
        record["program_name_normalized"] = normalize_program_name(
            record["performance_program_name"]
        )
        record["planned_target_numeric"] = _number(values[positions["목표치"]])
        record["actual_value_numeric"] = _number(values[positions["실적치"]])
        record["official_achievement_rate_numeric"] = _number(values[positions["달성률"]])
        record["actual_value_missing_flag"] = record["actual_value_raw"] is None
        rows.append(record)
    workbook.close()

    if not rows:
        raise ManualPerformanceError("골드셋에 비어 있지 않은 행이 없습니다.")
    return pd.DataFrame(rows).convert_dtypes()


def _validate_financial(financial: pd.DataFrame) -> pd.DataFrame:
    missing = REQUIRED_FINANCIAL_COLUMNS - set(financial.columns)
    if missing:
        raise ManualPerformanceError(
            f"재정 프로그램-연도 필수 열이 없습니다: {', '.join(sorted(missing))}"
        )
    result = financial.copy()
    result["ministry_code"] = result["ministry_code"].astype("string").str.zfill(3)
    result["program_code"] = result["program_code"].astype("string")
    result["fiscal_year"] = pd.to_numeric(result["fiscal_year"], errors="coerce").astype("Int64")
    result["program_name_normalized"] = result["program_name"].map(normalize_program_name)
    duplicate = result.duplicated(
        [
            "ministry_code",
            "fiscal_year",
            "field_name",
            "sector_name",
            "program_code",
            "program_name_normalized",
        ],
        keep=False,
    )
    if duplicate.any():
        raise ManualPerformanceError(
            f"재정 프로그램-연도 기본키가 중복되었습니다: {int(duplicate.sum())}행"
        )
    return result


def _indicator_status_counts(series: pd.Series) -> str:
    counts = Counter(str(value) for value in series.dropna().astype(str) if str(value).strip())
    return json.dumps(dict(sorted(counts.items())), ensure_ascii=False)


def _program_year_performance(indicators: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    keys = [
        "ministry_name",
        "fiscal_year",
        "program_goal_number",
        "program_name_normalized",
    ]
    for key, part in indicators.groupby(keys, dropna=False, sort=True):
        ministry_name, fiscal_year, program_goal_number, program_name_normalized = key
        performance_program_name = (
            part["performance_program_name"].dropna().astype(str).mode().iloc[0]
        )
        rows.append(
            {
                "ministry_name": ministry_name,
                "fiscal_year": fiscal_year,
                "program_goal_number": program_goal_number,
                "performance_program_name": performance_program_name,
                "source_program_code": part["source_program_code"].dropna().iloc[0]
                if part["source_program_code"].notna().any()
                else pd.NA,
                "source_field_name": part["source_field_name"].dropna().iloc[0]
                if "source_field_name" in part and part["source_field_name"].notna().any()
                else pd.NA,
                "source_sector_name": part["source_sector_name"].dropna().iloc[0]
                if "source_sector_name" in part and part["source_sector_name"].notna().any()
                else pd.NA,
                "program_mapping_status": part["program_mapping_status"].dropna().iloc[0]
                if "program_mapping_status" in part and part["program_mapping_status"].notna().any()
                else pd.NA,
                "program_name_normalized": program_name_normalized,
                "indicator_count": len(part),
                "reported_indicator_count": int(part["actual_value_raw"].notna().sum()),
                "achievement_rate_numeric_count": int(
                    part["official_achievement_rate_numeric"].notna().sum()
                ),
                "plan_report_mismatch_count": int(
                    part["plan_report_match_status_raw"].eq("불일치").sum()
                ),
                "plan_report_match_status_counts": _indicator_status_counts(
                    part["plan_report_match_status_raw"]
                ),
                "source_indicator_ids": json.dumps(
                    part["source_indicator_id"].astype(str).tolist(),
                    ensure_ascii=False,
                ),
            }
        )
    return pd.DataFrame(rows).convert_dtypes()


def match_program_year(
    program_year: pd.DataFrame,
    financial: pd.DataFrame,
    *,
    ministry_code: str,
) -> pd.DataFrame:
    """코드, 정확 명칭, 공백 정규화 명칭 순으로 유일한 경우만 자동 연결합니다."""
    fiscal = _validate_financial(financial)
    fiscal = fiscal.loc[fiscal["ministry_code"].eq(ministry_code.zfill(3))].copy()
    output_rows: list[dict[str, Any]] = []
    financial_payload_columns = [
        column
        for column in fiscal.columns
        if column
        not in {
            "ministry_code",
            "ministry_name",
            "fiscal_year",
            "program_code",
            "program_name",
            "program_name_normalized",
        }
    ]

    for row in program_year.to_dict(orient="records"):
        year_candidates = fiscal.loc[fiscal["fiscal_year"].eq(row["fiscal_year"])]
        structural_deletion = row.get("program_mapping_status") == "DELETED_TRANSFERRED"
        external_ministry = row.get("program_mapping_status") == "EXTERNAL_MINISTRY"
        if structural_deletion:
            candidates = year_candidates.iloc[0:0]
            match_status = "STRUCTURAL_PROGRAM_DELETED_TRANSFERRED"
        elif external_ministry:
            candidates = year_candidates.iloc[0:0]
            match_status = "EXTERNAL_MINISTRY_FINANCIAL_PROGRAM"
        else:
            source_code = _text(row.get("source_program_code"))
            if source_code:
                candidates = year_candidates.loc[year_candidates["program_code"].eq(source_code)]
                match_status = "EXACT_CODE"
                source_field = _text(row.get("source_field_name"))
                source_sector = _text(row.get("source_sector_name"))
                if source_field:
                    candidates = candidates.loc[candidates["field_name"].eq(source_field)]
                if source_sector:
                    candidates = candidates.loc[candidates["sector_name"].eq(source_sector)]
                if source_field or source_sector:
                    match_status = "EXACT_CONFIRMED_HIERARCHY"
                if len(candidates) > 1:
                    same_name = candidates.loc[
                        candidates["program_name_normalized"].eq(row["program_name_normalized"])
                    ]
                    if not same_name.empty:
                        candidates = same_name
            else:
                candidates = year_candidates.loc[
                    year_candidates["program_name"]
                    .astype(str)
                    .str.strip()
                    .eq(str(row["performance_program_name"]).strip())
                ]
                match_status = "EXACT_NAME"
                if candidates.empty:
                    candidates = year_candidates.loc[
                        year_candidates["program_name_normalized"].eq(
                            row["program_name_normalized"]
                        )
                    ]
                    match_status = "NORMALIZED_NAME"

        if len(candidates) > 1:
            usable = candidates.loc[
                candidates["program_code"].ne("UNKNOWN")
                & candidates["financial_linkage_status"].eq("COMPLETE")
            ]
            if len(usable) == 1:
                candidates = usable
                match_status = "EXACT_NAME_UNIQUE_FINANCIAL"

        matched = len(candidates) == 1
        if candidates.empty and not (structural_deletion or external_ministry):
            match_status = "MANUAL_REVIEW_NO_MATCH"
        elif len(candidates) > 1:
            match_status = "MANUAL_REVIEW_MULTIPLE_MATCHES"

        result = dict(row)
        result["ministry_code"] = ministry_code.zfill(3)
        result["program_match_status"] = match_status
        result["program_match_eligible"] = matched
        result["program_code"] = pd.NA
        result["financial_program_name"] = pd.NA
        for column in financial_payload_columns:
            result[column] = pd.NA
        if matched:
            fiscal_row = candidates.iloc[0]
            result["program_code"] = fiscal_row["program_code"]
            result["financial_program_name"] = fiscal_row["program_name"]
            for column in financial_payload_columns:
                result[column] = fiscal_row[column]
        result["performance_outcome_analysis_eligible"] = result["reported_indicator_count"] > 0
        result["same_year_financial_analysis_eligible"] = bool(
            matched and result.get("financial_linkage_status") == "COMPLETE"
        )
        result["same_year_joint_analysis_eligible"] = bool(
            result["performance_outcome_analysis_eligible"]
            and result["same_year_financial_analysis_eligible"]
        )
        output_rows.append(result)
    return pd.DataFrame(output_rows).convert_dtypes()


def _issues(
    manual_rows: pd.DataFrame,
    indicators: pd.DataFrame,
    program_year: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for row in manual_rows.loc[manual_rows["is_example"]].itertuples():
        rows.append(
            {
                "issue_type": "EXAMPLE_ROW",
                "severity": "INFO",
                "source_row_number": row.source_row_number,
                "source_indicator_id": row.source_indicator_id,
                "fiscal_year": row.fiscal_year,
                "performance_program_name": row.performance_program_name,
                "details": "예시행은 보존하되 분석대상에서 제외",
            }
        )
    for row in manual_rows.loc[
        ~manual_rows["is_example"] & ~manual_rows["analysis_eligible"]
    ].itertuples():
        rows.append(
            {
                "issue_type": "INPUT_ROW_INELIGIBLE",
                "severity": "BLOCKING",
                "source_row_number": row.source_row_number,
                "source_indicator_id": row.source_indicator_id,
                "fiscal_year": row.fiscal_year,
                "performance_program_name": row.performance_program_name,
                "details": "부처·연도 범위 또는 필수키 확인 필요",
            }
        )
    for row in indicators.loc[indicators["actual_value_missing_flag"]].itertuples():
        rows.append(
            {
                "issue_type": "ACTUAL_VALUE_MISSING",
                "severity": "ANALYSIS_LIMITATION",
                "source_row_number": row.source_row_number,
                "source_indicator_id": row.source_indicator_id,
                "fiscal_year": row.fiscal_year,
                "performance_program_name": row.performance_program_name,
                "details": "성과실적 결측; 프로그램 매칭에는 포함하되 성과분석 제한",
            }
        )
    for row in program_year.loc[~program_year["program_match_eligible"]].itertuples():
        rows.append(
            {
                "issue_type": row.program_match_status,
                "severity": "MANUAL_REVIEW",
                "source_row_number": pd.NA,
                "source_indicator_id": pd.NA,
                "fiscal_year": row.fiscal_year,
                "performance_program_name": row.performance_program_name,
                "details": f"영향 성과지표 {row.indicator_count}행; 프로그램코드 수동 확인",
            }
        )
    return pd.DataFrame(
        rows,
        columns=[
            "issue_type",
            "severity",
            "source_row_number",
            "source_indicator_id",
            "fiscal_year",
            "performance_program_name",
            "details",
        ],
    ).convert_dtypes()


def _data_dictionary() -> pd.DataFrame:
    rows = [
        ("source_indicator_id", "원본 골드셋 행ID", "string"),
        ("fiscal_year", "성과 대상 회계연도", "integer"),
        ("performance_program_name", "성과문서의 프로그램명", "string"),
        ("indicator_name_plan", "성과계획서 지표명", "string"),
        ("indicator_name_report", "성과보고서 지표명", "string"),
        ("planned_target_raw", "계획서 목표치 원문값", "string"),
        ("actual_value_raw", "보고서 실적치 원문값", "string"),
        (
            "official_achievement_rate_raw",
            "성과보고서 공식 달성률 원문값",
            "string",
        ),
        ("program_match_status", "재정 프로그램 매칭상태", "string"),
        ("program_code", "유일 매칭된 열린재정 프로그램코드", "string"),
        ("original_budget", "프로그램 확정 본예산 합계", "integer"),
        ("current_budget", "프로그램 예산현액 합계", "integer"),
        ("settlement_expenditure", "프로그램 결산 지출액 합계", "integer"),
        ("execution_rate", "회계유형별 확인된 분자·분모로 계산한 집행률", "float"),
        (
            "same_year_joint_analysis_eligible",
            "성과실적과 완전 재정연결을 함께 사용할 수 있는지 여부",
            "boolean",
        ),
    ]
    return pd.DataFrame(rows, columns=["column", "definition", "logical_type"])


def build_manual_performance_pilot(
    *,
    input_path: Path,
    financial_path: Path,
    output_dir: Path,
    ministry_name: str = "중소벤처기업부",
    ministry_code: str = "102",
    start_year: int = 2022,
    end_year: int = 2024,
    program_code_confirmations_path: Path | None = None,
    overwrite: bool = False,
) -> ManualPerformanceResult:
    output_paths = [
        output_dir / "manual_performance_rows.parquet",
        output_dir / "program_kpi_year.parquet",
        output_dir / "program_year_performance_financial.parquet",
        output_dir / "manual_review.csv",
        output_dir / "data_dictionary.csv",
        output_dir / "normalization_summary.json",
    ]
    existing = [path for path in output_paths if path.exists()]
    if existing and not overwrite:
        raise FileExistsError(
            "기존 산출물이 있습니다. --overwrite를 지정하세요: "
            + ", ".join(str(path) for path in existing)
        )
    if not financial_path.is_file():
        raise ManualPerformanceError(
            f"프로그램-연도 재정 마스터를 찾을 수 없습니다: {financial_path}"
        )

    source_hash_before = _source_sha256(input_path)
    manual_rows = read_manual_performance_workbook(
        input_path,
        ministry_name=ministry_name,
        start_year=start_year,
        end_year=end_year,
    )
    indicators = manual_rows.loc[manual_rows["analysis_eligible"]].copy()
    indicators = apply_program_code_confirmations(
        indicators,
        program_code_confirmations_path,
    )
    duplicate_indicator_ids = indicators.duplicated("source_indicator_id", keep=False)
    if duplicate_indicator_ids.any():
        raise ManualPerformanceError(
            f"분석대상 성과지표 행ID가 중복되었습니다: {int(duplicate_indicator_ids.sum())}행"
        )

    financial = pd.read_parquet(financial_path)
    program_performance = _program_year_performance(indicators)
    program_year = match_program_year(
        program_performance,
        financial,
        ministry_code=ministry_code,
    )
    issues = _issues(manual_rows, indicators, program_year)
    data_dictionary = _data_dictionary()

    matched = program_year["program_match_eligible"].astype(bool)
    amount_columns = ["original_budget", "current_budget", "settlement_expenditure"]
    amount_reconciliation: dict[str, dict[str, int]] = {}
    fiscal_checked = _validate_financial(financial)
    fiscal_checked = fiscal_checked.loc[
        fiscal_checked["ministry_code"].eq(ministry_code.zfill(3))
    ].copy()
    for column in amount_columns:
        match_key = [
            "fiscal_year",
            "field_name",
            "sector_name",
            "program_code",
            "program_name",
        ]
        matched_programs = (
            program_year.loc[
                matched,
                [
                    "fiscal_year",
                    "field_name",
                    "sector_name",
                    "program_code",
                    "financial_program_name",
                    column,
                ],
            ]
            .rename(columns={"financial_program_name": "program_name"})
            .drop_duplicates(match_key)
        )
        output_amount = int(
            pd.to_numeric(matched_programs[column], errors="coerce").sum(skipna=True)
        )
        matched_keys = matched_programs.loc[:, match_key]
        source_subset = fiscal_checked.merge(
            matched_keys,
            how="inner",
            on=match_key,
            validate="one_to_one",
        )
        source_amount = int(pd.to_numeric(source_subset[column], errors="coerce").sum(skipna=True))
        amount_reconciliation[column] = {
            "source_amount": source_amount,
            "output_amount": output_amount,
            "difference": output_amount - source_amount,
        }

    summary = {
        "generated_at": datetime.now(UTC).isoformat(),
        "pilot_scope": {
            "ministry_name": ministry_name,
            "ministry_code": ministry_code.zfill(3),
            "start_year": start_year,
            "end_year": end_year,
        },
        "source_file": str(input_path),
        "source_sha256": source_hash_before,
        "raw_nonempty_row_count": len(manual_rows),
        "example_row_count": int(manual_rows["is_example"].sum()),
        "analysis_indicator_row_count": len(indicators),
        "indicator_year_counts": {
            str(key): int(value)
            for key, value in indicators["fiscal_year"].value_counts().sort_index().items()
        },
        "indicator_id_duplicate_count": int(indicators.duplicated("source_indicator_id").sum()),
        "actual_value_missing_count": int(indicators["actual_value_missing_flag"].sum()),
        "program_year_count": len(program_year),
        "program_match_status_counts": {
            str(key): int(value)
            for key, value in program_year["program_match_status"].value_counts().items()
        },
        "matched_program_year_count": int(matched.sum()),
        "manual_review_program_year_count": int((~matched).sum()),
        "same_year_joint_analysis_eligible_count": int(
            program_year["same_year_joint_analysis_eligible"].sum()
        ),
        "financial_linkage_status_counts": {
            str(key): int(value)
            for key, value in program_year.loc[matched, "financial_linkage_status"]
            .value_counts(dropna=False)
            .items()
        },
        "amount_reconciliation": amount_reconciliation,
        "issue_type_counts": {
            str(key): int(value) for key, value in issues["issue_type"].value_counts().items()
        },
        "validation": {
            "raw_rows_partitioned": len(manual_rows)
            == len(indicators) + int((~manual_rows["analysis_eligible"]).sum()),
            "indicator_id_unique": not indicators["source_indicator_id"].duplicated().any(),
            "program_year_key_unique": not program_year.duplicated(
                [
                    "ministry_code",
                    "fiscal_year",
                    "program_goal_number",
                    "program_name_normalized",
                ]
            ).any(),
            "all_auto_matches_unique": bool(
                program_year.loc[matched, "program_code"].notna().all()
            ),
            "amounts_preserved": all(
                values["difference"] == 0 for values in amount_reconciliation.values()
            ),
            "source_file_unchanged": source_hash_before == _source_sha256(input_path),
        },
        "interpretation_limit": (
            "성과지표를 세부사업에 귀속하지 않으며, 복수 지표의 달성률을 평균하지 "
            "않습니다. 미매칭 프로그램과 재정 PARTIAL/UNMATCHED는 공동분석에서 제한합니다."
        ),
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    manual_rows.to_parquet(output_paths[0], index=False)
    indicators.to_parquet(output_paths[1], index=False)
    program_year.to_parquet(output_paths[2], index=False)
    issues.to_csv(output_paths[3], index=False, encoding="utf-8-sig")
    data_dictionary.to_csv(output_paths[4], index=False, encoding="utf-8-sig")
    output_paths[5].write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return ManualPerformanceResult(
        manual_rows=manual_rows,
        indicators=indicators,
        program_year=program_year,
        issues=issues,
        data_dictionary=data_dictionary,
        summary=summary,
        output_paths=output_paths,
    )


def build_program_match_review(
    *,
    program_year_path: Path,
    financial_path: Path,
    output_dir: Path,
    ministry_code: str,
    candidate_count: int = 3,
    overwrite: bool = False,
) -> tuple[pd.DataFrame, dict[str, Any], tuple[Path, ...]]:
    """미매칭 프로그램마다 같은 부처·연도의 명칭 유사 후보를 제시합니다."""
    if candidate_count < 1:
        raise ManualPerformanceError("후보 수는 1개 이상이어야 합니다.")
    output_paths = (
        output_dir / "program_match_candidates.csv",
        output_dir / "program_match_review_summary.json",
    )
    existing = [path for path in output_paths if path.exists()]
    if existing and not overwrite:
        raise FileExistsError(
            "기존 산출물이 있습니다. --overwrite를 지정하세요: "
            + ", ".join(str(path) for path in existing)
        )
    for path in (program_year_path, financial_path):
        if not path.is_file():
            raise ManualPerformanceError(f"입력 파일을 찾을 수 없습니다: {path}")

    source_hashes = {
        str(path): _source_sha256(path) for path in (program_year_path, financial_path)
    }
    program_year = pd.read_parquet(program_year_path)
    financial = _validate_financial(pd.read_parquet(financial_path))
    required = {
        "fiscal_year",
        "performance_program_name",
        "indicator_count",
        "source_indicator_ids",
        "program_match_status",
        "program_match_eligible",
    }
    missing = sorted(required - set(program_year.columns))
    if missing:
        raise ManualPerformanceError(f"프로그램-연도 입력에 필수 컬럼이 없습니다: {missing}")

    code = ministry_code.zfill(3)
    review = program_year.loc[~program_year["program_match_eligible"].fillna(False)].copy()
    candidates = financial.loc[
        financial["ministry_code"].eq(code) & financial["program_code"].ne("UNKNOWN")
    ].copy()
    rows: list[dict[str, Any]] = []
    for source in review.sort_values(["fiscal_year", "performance_program_name"]).itertuples():
        same_year = candidates.loc[candidates["fiscal_year"].eq(source.fiscal_year)].copy()
        normalized = normalize_program_name(source.performance_program_name)
        same_year["name_similarity"] = same_year["program_name"].map(
            lambda name, target=normalized: SequenceMatcher(
                None,
                target,
                normalize_program_name(name),
            ).ratio()
        )
        same_year = same_year.sort_values(
            ["name_similarity", "original_budget", "program_code"],
            ascending=[False, False, True],
        ).head(candidate_count)
        for rank, candidate in enumerate(same_year.itertuples(), start=1):
            score = float(candidate.name_similarity)
            rows.append(
                {
                    "review_key": (
                        f"{code}|{source.fiscal_year}|{source.performance_program_name}"
                    ),
                    "ministry_code": code,
                    "fiscal_year": int(source.fiscal_year),
                    "performance_program_name": source.performance_program_name,
                    "review_group": normalized,
                    "indicator_count": int(source.indicator_count),
                    "source_indicator_ids": source.source_indicator_ids,
                    "source_match_status": source.program_match_status,
                    "candidate_rank": rank,
                    "candidate_fiscal_year": int(candidate.fiscal_year),
                    "candidate_program_code": str(candidate.program_code),
                    "candidate_program_name": candidate.program_name,
                    "name_similarity": score,
                    "candidate_quality": (
                        "HIGH" if score >= 0.8 else "MEDIUM" if score >= 0.6 else "LOW"
                    ),
                    "candidate_original_budget": candidate.original_budget,
                    "candidate_financial_linkage_status": candidate.financial_linkage_status,
                    "auto_confirmed": False,
                    "decision": "",
                    "reviewer": "",
                    "review_note": "",
                }
            )
    result = pd.DataFrame(rows)
    if len(review) and result.empty:
        raise ManualPerformanceError("미매칭 프로그램의 같은 연도 재정 후보가 없습니다.")
    if not result.empty:
        if result["review_key"].nunique() != len(review):
            raise ManualPerformanceError("일부 미매칭 프로그램에 후보가 생성되지 않았습니다.")
        if result["auto_confirmed"].any():
            raise ManualPerformanceError("후보를 자동 확정해서는 안 됩니다.")
        if not result["ministry_code"].eq(code).all():
            raise ManualPerformanceError("다른 부처 후보가 섞였습니다.")

    source_hashes_after = {
        str(path): _source_sha256(path) for path in (program_year_path, financial_path)
    }
    top = result.loc[result["candidate_rank"].eq(1)]
    summary = {
        "generated_at": datetime.now(UTC).isoformat(),
        "ministry_code": code,
        "source_unmatched_program_year_count": len(review),
        "source_review_group_count": int(result["review_group"].nunique()),
        "candidate_row_count": len(result),
        "candidate_count_per_source": candidate_count,
        "top_candidate_quality_counts": {
            str(key): int(value) for key, value in top["candidate_quality"].value_counts().items()
        },
        "top_candidate_similarity": {
            "min": float(top["name_similarity"].min()) if len(top) else None,
            "median": float(top["name_similarity"].median()) if len(top) else None,
            "max": float(top["name_similarity"].max()) if len(top) else None,
        },
        "source_sha256": source_hashes,
        "validation": {
            "all_source_rows_preserved": result["review_key"].nunique() == len(review),
            "same_ministry_only": bool(result["ministry_code"].eq(code).all()),
            "same_year_only": bool(result["fiscal_year"].eq(result["candidate_fiscal_year"]).all()),
            "auto_confirmed_count": int(result["auto_confirmed"].sum()),
            "input_files_unchanged": source_hashes == source_hashes_after,
        },
        "interpretation_limit": (
            "명칭 유사도는 검토 순서를 위한 후보일 뿐 매칭 근거가 아닙니다. "
            "프로그램 코드와 공식 문서 근거를 사람이 확인한 뒤 결정해야 합니다."
        ),
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    result.to_csv(output_paths[0], index=False, encoding="utf-8-sig")
    output_paths[1].write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return result, summary, output_paths
