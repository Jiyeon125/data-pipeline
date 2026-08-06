"""Build the human-review workbook and convert completed responses to stage-1 CSV."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SOURCE_CSV = Path("validation/blind_pair_stage1_anonymous.csv")
WORKBOOK_PATH = Path("validation/blind_pair_human_review.xlsx")
OUTPUT_CSV = Path("validation/blind_pair_stage1_review.csv")
HEADER_ROW = 7
DATA_START_ROW = 8

CHOICES = ("CASE_A", "CASE_B", "BOTH", "NEITHER", "UNABLE_TO_DECIDE")
CONFIDENCE_LEVELS = ("LOW", "MEDIUM", "HIGH")
TARGET_STATUS_LABELS = {
    "ALL_COMPARABLE_AT_OR_ABOVE_TARGET": "비교 가능 지표 모두 목표 달성 이상",
    "ALL_COMPARABLE_BELOW_TARGET": "비교 가능 지표 모두 목표 미달",
    "MIXED_COMPARABLE": "비교 가능 지표의 달성·미달이 혼합됨",
    "NO_COMPARABLE_RATE": "비교 가능한 달성률 없음",
}
SIZE_LABELS = {
    "SMALL": "100억 원 미만",
    "MEDIUM": "100억~1,000억 원 미만",
    "LARGE": "1,000억~1조 원 미만",
    "VERY_LARGE": "1조 원 이상",
    "UNKNOWN": "규모 확인 불가",
}
QUALITY_LABELS = {"ANALYZABLE": "분석 가능", "DATA_QUALITY_HOLD": "데이터 품질 확인 필요"}
SOURCE_LABELS = {"SOURCE_LOCATION_NOT_ATTACHED": "원문 위치 미첨부"}

CASE_FIELDS = (
    ("익명 사례 ID", "anonymous_case_id"),
    ("회계연도", "fiscal_year"),
    ("익명 부처", "ministry_code_anon"),
    ("익명 프로그램", "program_code_anon"),
    ("예산 규모", "budget_size_band"),
    ("본예산(억원)", "original_budget"),
    ("예산현액(억원)", "current_budget"),
    ("지출액(억원)", "expenditure"),
    ("집행률", "execution_rate"),
    ("보고목표 상태", "reported_target_status"),
    ("반복 관측", "repeated_observation_summary"),
    ("예산 변화", "budget_change_summary"),
    ("데이터 품질", "data_quality_status"),
    ("원문 확인", "source_review_availability"),
)
REVIEW_HEADERS = (
    "검토자",
    "먼저 확인할 사례",
    "선택 이유",
    "둘 다 우선인 이유",
    "둘 다 후순위인 이유",
    "판단 불가 이유",
    "추가 필요자료",
    "확신도",
    "검토일시",
)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def _bool_word(value: str) -> str:
    return "있음" if value.strip().lower() == "true" else "없음"


def _repeated_summary(value: str) -> str:
    parsed = dict(part.split("=", 1) for part in value.split(";") if "=" in part)
    return " · ".join(
        (
            f"저집행 관측 {_bool_word(parsed.get('low_execution', 'false'))}",
            f"연속 목표미달 {_bool_word(parsed.get('target_miss_consecutive', 'false'))}",
            f"연속 목표달성 {_bool_word(parsed.get('target_met_consecutive', 'false'))}",
        )
    )


def _display_value(field: str, value: str) -> object:
    if field in {"original_budget", "current_budget", "expenditure"}:
        return None if not value else float(value) / 100_000_000
    if field == "execution_rate":
        return None if not value else float(value)
    if field == "budget_size_band":
        return SIZE_LABELS.get(value, value)
    if field == "reported_target_status":
        return TARGET_STATUS_LABELS.get(value, "보고목표 상태 확인 필요")
    if field == "repeated_observation_summary":
        return _repeated_summary(value)
    if field == "budget_change_summary":
        rate = value.removeprefix("change_rate=")
        return "비교 불가" if not rate or rate == "NA" else float(rate)
    if field == "data_quality_status":
        return QUALITY_LABELS.get(value, value)
    if field == "source_review_availability":
        return SOURCE_LABELS.get(value, value)
    if field == "fiscal_year":
        return int(value)
    return value


def build_workbook(source_csv: Path, output_xlsx: Path) -> None:
    _, rows = read_csv(source_csv)
    pairs: dict[str, dict[str, dict[str, str]]] = {}
    for row in rows:
        pairs.setdefault(row["pair_id"], {})[row["case_label"]] = row
    if any(set(cases) != {"CASE_A", "CASE_B"} for cases in pairs.values()):
        raise ValueError("각 pair_id에는 CASE_A와 CASE_B가 각각 한 행씩 있어야 합니다.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "사람검토"
    headers = (
        ["비교쌍 ID"]
        + [
            f"사례 {label[-1]} - {korean}"
            for label in ("CASE_A", "CASE_B")
            for korean, _ in CASE_FIELDS
        ]
        + list(REVIEW_HEADERS)
    )

    instructions = (
        "블라인드 쌍대비교 작성방법",
        "두 사례 중 원문을 먼저 확인할 사례를 선택합니다. 사업의 성과등급을 매기는 작업이 아닙니다.",
        "외부검색 없이 제공된 사실만 보고 작성하며, 판단할 수 없으면 UNABLE_TO_DECIDE를 선택합니다.",
        "CASE_A·CASE_B·BOTH·NEITHER·UNABLE_TO_DECIDE 중 하나와 선택 이유를 작성합니다.",
        "확신도는 LOW·MEDIUM·HIGH 중 하나를 선택하고 검토일시를 입력합니다.",
    )
    for row_number, text in enumerate(instructions, 1):
        sheet.merge_cells(
            start_row=row_number, start_column=1, end_row=row_number, end_column=len(headers)
        )
        cell = sheet.cell(row_number, 1, text)
        cell.font = Font(bold=row_number == 1, size=14 if row_number == 1 else 10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row_number].height = 24 if row_number == 1 else 20

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(HEADER_ROW, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        if header.startswith("사례 A"):
            cell.fill = PatternFill("solid", fgColor="4472C4")
        elif header.startswith("사례 B"):
            cell.fill = PatternFill("solid", fgColor="548235")
        elif header in REVIEW_HEADERS:
            cell.fill = PatternFill("solid", fgColor="BF9000")
        else:
            cell.fill = PatternFill("solid", fgColor="5B6573")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for output_row, pair_id in enumerate(sorted(pairs), DATA_START_ROW):
        values: list[object] = [pair_id]
        for label in ("CASE_A", "CASE_B"):
            case = pairs[pair_id][label]
            values.extend(_display_value(field, case[field]) for _, field in CASE_FIELDS)
        values.extend([None] * len(REVIEW_HEADERS))
        for column, value in enumerate(values, 1):
            sheet.cell(output_row, column, value)

    money_headers = {
        f"사례 {label} - {name}"
        for label in ("A", "B")
        for name in ("본예산(억원)", "예산현액(억원)", "지출액(억원)")
    }
    percent_headers = {
        f"사례 {label} - {name}" for label in ("A", "B") for name in ("집행률", "예산 변화")
    }
    header_columns = {cell.value: cell.column for cell in sheet[HEADER_ROW]}
    for header in money_headers:
        for row in range(DATA_START_ROW, DATA_START_ROW + len(pairs)):
            sheet.cell(row, header_columns[header]).number_format = "#,##0.0"
    for header in percent_headers:
        for row in range(DATA_START_ROW, DATA_START_ROW + len(pairs)):
            sheet.cell(row, header_columns[header]).number_format = "0.0%"

    choice_validation = DataValidation(type="list", formula1='"' + ",".join(CHOICES) + '"')
    confidence_validation = DataValidation(
        type="list", formula1='"' + ",".join(CONFIDENCE_LEVELS) + '"'
    )
    sheet.add_data_validation(choice_validation)
    sheet.add_data_validation(confidence_validation)
    last_row = DATA_START_ROW + len(pairs) - 1
    choice_validation.add(
        f"{get_column_letter(header_columns['먼저 확인할 사례'])}{DATA_START_ROW}:"
        f"{get_column_letter(header_columns['먼저 확인할 사례'])}{last_row}"
    )
    confidence_validation.add(
        f"{get_column_letter(header_columns['확신도'])}{DATA_START_ROW}:"
        f"{get_column_letter(header_columns['확신도'])}{last_row}"
    )

    thin = Side(style="thin", color="D9D9D9")
    for row in sheet.iter_rows(min_row=HEADER_ROW, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, header in enumerate(headers, 1):
        width = 14
        if any(
            word in header for word in ("보고목표", "반복 관측", "선택 이유", "이유", "필요자료")
        ):
            width = 24
        elif "익명 사례" in header or "원문 확인" in header:
            width = 18
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = f"B{DATA_START_ROW}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[HEADER_ROW].height = 42
    for row in range(DATA_START_ROW, last_row + 1):
        sheet.row_dimensions[row].height = 72
    review_start = header_columns["검토자"]
    review_end = header_columns["검토일시"]
    sheet.conditional_formatting.add(
        f"{get_column_letter(review_start)}{DATA_START_ROW}:{get_column_letter(review_end)}{last_row}",
        FormulaRule(
            formula=[f"LEN(${get_column_letter(review_start)}{DATA_START_ROW})>0"],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)


def convert_responses(input_xlsx: Path, source_csv: Path, output_csv: Path) -> None:
    workbook = load_workbook(input_xlsx, data_only=False)
    sheet = workbook.worksheets[0]
    header_row = next(
        (
            row
            for row in range(1, min(sheet.max_row, 20) + 1)
            if sheet.cell(row, 1).value == "비교쌍 ID"
        ),
        None,
    )
    if header_row is None:
        raise ValueError("'비교쌍 ID' 헤더를 찾지 못했습니다.")
    columns = {str(cell.value): cell.column for cell in sheet[header_row] if cell.value}
    required = {"비교쌍 ID", *REVIEW_HEADERS}
    if missing := required - set(columns):
        raise ValueError(f"검토 응답 컬럼이 없습니다: {sorted(missing)}")

    responses: dict[str, dict[str, str]] = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        pair_id = str(sheet.cell(row, columns["비교쌍 ID"]).value or "").strip()
        if not pair_id:
            continue
        response = {
            header: str(sheet.cell(row, columns[header]).value or "").strip()
            for header in REVIEW_HEADERS
        }
        if response["먼저 확인할 사례"] not in CHOICES:
            raise ValueError(f"{pair_id}: 먼저 확인할 사례를 드롭다운 값으로 입력하세요.")
        if response["확신도"] not in CONFIDENCE_LEVELS:
            raise ValueError(f"{pair_id}: 확신도를 드롭다운 값으로 입력하세요.")
        if not response["검토자"]:
            raise ValueError(f"{pair_id}: 검토자를 입력하세요.")
        responses[pair_id] = response

    fieldnames, source_rows = read_csv(source_csv)
    source_pairs = {row["pair_id"] for row in source_rows}
    if set(responses) != source_pairs:
        missing = sorted(source_pairs - set(responses))
        extra = sorted(set(responses) - source_pairs)
        raise ValueError(f"응답 pair_id 불일치: missing={missing}, extra={extra}")
    for row in source_rows:
        response = responses[row["pair_id"]]
        row.update(
            {
                "reviewer_id": response["검토자"],
                "first_review_choice": response["먼저 확인할 사례"],
                "decision": "",
                "decision_reason": response["선택 이유"],
                "both_priority_reason": response["둘 다 우선인 이유"],
                "both_low_priority_reason": response["둘 다 후순위인 이유"],
                "unable_to_decide_reason": response["판단 불가 이유"],
                "additional_information_needed": response["추가 필요자료"],
                "confidence_level": response["확신도"],
                "reviewed_at": response["검토일시"],
            }
        )
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(source_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build")
    build.add_argument("--source", type=Path, default=SOURCE_CSV)
    build.add_argument("--output", type=Path, default=WORKBOOK_PATH)
    convert = subparsers.add_parser("convert")
    convert.add_argument("--input", type=Path, default=WORKBOOK_PATH)
    convert.add_argument("--source", type=Path, default=SOURCE_CSV)
    convert.add_argument("--output", type=Path, default=OUTPUT_CSV)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.command == "build":
        build_workbook(args.source, args.output)
    else:
        convert_responses(args.input, args.source, args.output)


if __name__ == "__main__":
    main()
