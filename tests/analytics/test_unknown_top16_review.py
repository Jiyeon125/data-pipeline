from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from analytics.unknown_top16_review import (
    HEADER_ROW,
    PROJECT_SOURCE_HEADERS,
    SHEET_PROJECTS,
    UnknownReviewError,
    UnknownReviewPaths,
    apply_unknown_review_overlay,
    build_unknown_review_workbook,
    validate_unknown_review_workbook,
)


def _paths(tmp_path: Path) -> UnknownReviewPaths:
    return UnknownReviewPaths(
        priority=tmp_path / "unknown.csv",
        ranking_population=tmp_path / "ranking.parquet",
        workbook=tmp_path / "review.xlsx",
        validation_summary=tmp_path / "validation.json",
    )


def _write_inputs(paths: UnknownReviewPaths) -> None:
    priority = pd.DataFrame(
        [
            {
                "classification_project_id": "classification-code:019:1:10:20:30",
                "ministry_code": "019",
                "budget_coverage_order": 1,
                "priority_80pct_coverage": True,
                "cumulative_unknown_budget_share": 0.8,
            }
        ]
    )
    priority.to_csv(paths.priority, index=False)
    rows = []
    for year, amount in zip(range(2023, 2026), [20, 30, 40], strict=True):
        rows.append(
            {
                "classification_project_id": "classification-code:019:1:10:20:30",
                "fiscal_year": year,
                "ministry_code": "019",
                "analysis_ministry_name": "고용노동부",
                "account_type_classified": "FUND",
                "program_code": "10",
                "program_name": "프로그램",
                "activity_code": "20",
                "activity_name": "단위사업",
                "subactivity_code": "30",
                "subactivity_name": "세부사업",
                "original_budget_analysis_amount": amount,
            }
        )
    pd.DataFrame(rows).to_parquet(paths.ranking_population, index=False)


def test_build_and_validate_incomplete_review_workbook(tmp_path: Path) -> None:
    paths = _paths(tmp_path)
    _write_inputs(paths)
    output = build_unknown_review_workbook(paths, expected_project_count=1)
    assert output.exists()

    workbook = load_workbook(output, data_only=False)
    assert workbook[SHEET_PROJECTS].max_row == 5
    assert workbook[SHEET_PROJECTS].cell(HEADER_ROW, 1).value == PROJECT_SOURCE_HEADERS[0]

    result = validate_unknown_review_workbook(paths, expected_project_count=1)
    assert result.status == "INCOMPLETE"
    assert result.project_count == 1
    assert result.year_row_count == 3
    assert result.confirmed_project_count == 0
    assert result.warning_count == 1


def test_confirmed_review_requires_evidence(tmp_path: Path) -> None:
    paths = _paths(tmp_path)
    _write_inputs(paths)
    build_unknown_review_workbook(paths, expected_project_count=1)
    workbook = load_workbook(paths.workbook)
    sheet = workbook[SHEET_PROJECTS]
    headers = [cell.value for cell in sheet[HEADER_ROW]]
    sheet.cell(5, headers.index("review_status") + 1, "CONFIRMED")
    sheet.cell(5, headers.index("analysis_scope_status") + 1, "IN_SCOPE")
    sheet.cell(5, headers.index("fiscal_instrument_applicability") + 1, "APPLICABLE")
    sheet.cell(5, headers.index("fiscal_instrument") + 1, "DIRECT")
    sheet.cell(5, headers.index("all_years_same_classification") + 1, "YES")
    workbook.save(paths.workbook)

    result = validate_unknown_review_workbook(paths, expected_project_count=1)
    assert result.status == "FAIL"
    assert any("CONFIRMED 필수값 누락" in error for error in result.errors)


def test_require_complete_fails_for_unreviewed_rows(tmp_path: Path) -> None:
    paths = _paths(tmp_path)
    _write_inputs(paths)
    build_unknown_review_workbook(paths, expected_project_count=1)
    result = validate_unknown_review_workbook(
        paths,
        require_complete=True,
        expected_project_count=1,
    )
    assert result.status == "FAIL"
    assert any("CONFIRMED가 필요" in error for error in result.errors)


def test_existing_review_workbook_is_not_overwritten_by_default(tmp_path: Path) -> None:
    paths = _paths(tmp_path)
    _write_inputs(paths)
    build_unknown_review_workbook(paths, expected_project_count=1)
    original = paths.workbook.read_bytes()

    with pytest.raises(UnknownReviewError, match="덮어쓰지 않습니다"):
        build_unknown_review_workbook(paths, expected_project_count=1)

    assert paths.workbook.read_bytes() == original


def test_confirmed_structural_mixed_review_stays_unknown(tmp_path: Path) -> None:
    paths = _paths(tmp_path)
    _write_inputs(paths)
    build_unknown_review_workbook(paths, expected_project_count=1)
    workbook = load_workbook(paths.workbook)
    sheet = workbook[SHEET_PROJECTS]
    headers = [cell.value for cell in sheet[HEADER_ROW]]
    values = {
        "analysis_scope_status": "IN_SCOPE",
        "fiscal_instrument_applicability": "REVIEW_REQUIRED",
        "fiscal_instrument": "UNKNOWN",
        "all_years_same_classification": "YES",
        "classification_evidence": "공식 자료에서 직접과 보조가 함께 확인됨",
        "evidence_source": "공식 사업설명자료 p.1",
        "confidence": "HIGH",
        "reviewer": "검수자",
        "reviewed_at": "2026-07-31",
        "review_status": "CONFIRMED",
        "review_note": "일반 분석 유지, 재정수단 내부 비교 제한",
    }
    for header, value in values.items():
        sheet.cell(5, headers.index(header) + 1, value)
    workbook.save(paths.workbook)

    validation = validate_unknown_review_workbook(
        paths,
        require_complete=True,
        expected_project_count=1,
    )
    frame = pd.read_parquet(paths.ranking_population).assign(
        fiscal_instrument="UNKNOWN",
        account_type_classified="FUND",
        project_category="PROGRAM_EXPENDITURE",
        analysis_included_classified=True,
        exclusion_category_classified=pd.NA,
        exclusion_reason_classified=pd.NA,
        instrument_classification_method="NO_INSTRUMENT_RULE_MATCH",
        instrument_classification_evidence="",
        instrument_manual_review_required=True,
        classification_manual_review_required=True,
        classification_status="MANUAL_REVIEW",
        classification_method="RULE",
        classification_evidence="",
        comparison_group="FUND|UNKNOWN|PROGRAM_EXPENDITURE",
    )
    overlaid, audit = apply_unknown_review_overlay(frame, paths.workbook)

    assert validation.status == "PASS"
    assert overlaid["fiscal_instrument"].eq("UNKNOWN").all()
    assert overlaid["analysis_included_classified"].all()
    assert overlaid["classification_status"].eq("MANUAL_CONFIRMED").all()
    assert audit["ranking_population_impact"].eq("KEEP_GENERAL_EXCLUDE_INSTRUMENT_PEER").all()
