import csv
from pathlib import Path

from openpyxl import load_workbook

from validation.convert_blind_pair_human_review import (
    CHOICES,
    CONFIDENCE_LEVELS,
    build_workbook,
    convert_responses,
)
from validation.summarize_blind_review import main


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def test_no_human_response_and_completed_summary(tmp_path: Path) -> None:
    stage1 = tmp_path / "reviewer.csv"
    stage2 = tmp_path / "questions.csv"
    key = tmp_path / "key.csv"
    summary = tmp_path / "summary.csv"
    report = tmp_path / "report.md"
    write_csv(
        stage1,
        ["pair_id", "reviewer_id", "first_review_choice", "additional_information_needed"],
        [{"pair_id": "BP01"}],
    )
    write_csv(stage2, ["pair_id", "case_label", "reviewer_question_rating"], [])
    write_csv(
        key,
        ["pair_id", "case_label", "review_grade", "selection_stratum"],
        [
            {
                "pair_id": "BP01",
                "case_label": "CASE_A",
                "review_grade": "A",
                "selection_stratum": "A 대 C",
            },
            {
                "pair_id": "BP01",
                "case_label": "CASE_B",
                "review_grade": "C",
                "selection_stratum": "A 대 C",
            },
        ],
    )
    args = [
        "--stage1",
        str(stage1),
        "--stage2",
        str(stage2),
        "--answer-key",
        str(key),
        "--summary-csv",
        str(summary),
        "--report",
        str(report),
    ]
    assert main(args) == 2
    assert not summary.exists()

    write_csv(
        stage1,
        ["pair_id", "reviewer_id", "first_review_choice", "additional_information_needed"],
        [
            {
                "pair_id": "BP01",
                "reviewer_id": "R1",
                "first_review_choice": "CASE_A",
                "additional_information_needed": "결산 원문",
            }
        ],
    )
    assert main(args) == 0
    assert "model_higher_grade_case_selected_count" in summary.read_text(encoding="utf-8-sig")
    assert "결산 원문" in report.read_text(encoding="utf-8")

    source = Path("validation/blind_pair_stage1_anonymous.csv")
    workbook_path = tmp_path / "human_review.xlsx"
    converted = tmp_path / "converted.csv"
    build_workbook(source, workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["사람검토"]
    headers = [cell.value for cell in sheet[7]]
    assert (sheet.max_row, sheet.max_column) == (15, 38)
    assert len(sheet.data_validations.dataValidation) == 2
    assert sheet.cell(8, headers.index("사례 A - 본예산(억원)") + 1).number_format == "#,##0.0"
    assert "MIXED_COMPARABLE" not in str(
        sheet.cell(8, headers.index("사례 A - 보고목표 상태") + 1).value
    )
    assert not any(
        forbidden in str(value)
        for row in sheet.iter_rows(values_only=True)
        for value in row
        for forbidden in ("review_grade", "diagnostic_type", "grade_reason_codes")
    )
    response_columns = {cell.value: cell.column for cell in sheet[7]}
    for row in range(8, 16):
        sheet.cell(row, response_columns["검토자"], "R1")
        sheet.cell(row, response_columns["먼저 확인할 사례"], CHOICES[0])
        sheet.cell(row, response_columns["선택 이유"], "원문 확인 필요")
        sheet.cell(row, response_columns["확신도"], CONFIDENCE_LEVELS[-1])
    workbook.save(workbook_path)
    convert_responses(workbook_path, source, converted)
    converted_rows = list(csv.DictReader(converted.open(encoding="utf-8-sig", newline="")))
    assert len(converted_rows) == 16
    assert {row["first_review_choice"] for row in converted_rows} == {"CASE_A"}
